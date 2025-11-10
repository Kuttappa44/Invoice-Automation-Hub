#!/usr/bin/env python3
"""
Final Invoice Processor - Complete Flow
1. Check unread emails
2. Process invoices with AWS Bedrock Vision
3. Upload PDFs to Google Drive folders
4. Update Excel sheet with extracted data
"""

import imaplib
import email
import re
import os
from datetime import datetime, timedelta
import pandas as pd
from fuzzywuzzy import fuzz, process
from dotenv import load_dotenv
import PyPDF2
from google_drive_uploader import GoogleDriveUploader
from openai_vision_extractor import OpenAIPropertyExtractor
import io

# Load environment variables
load_dotenv()

# Invoice keywords
INVOICE_KEYWORDS = [
    'invoice', 'bill', 'receipt', 'payment', 'booking', 'reservation',
    'confirmation', 'voucher', 'ticket', 'statement', 'charge',
    'hotel', 'accommodation', 'travel', 'booking id', 'reservation id',
    'pending', 'required', 'urgent', 'bills'
]

def load_vendor_reference():
    """Load the Excel reference sheet with vendor names"""
    try:
        df = pd.read_excel('userlist.xlsx')
        vendor_reference = {}
        columns = ['SANDHYA', 'MOKSHITHA', 'KUMAR', 'LAKSHMI']
        
        for col in columns:
            if col in df.columns:
                vendors = df[col].dropna().astype(str).tolist()
                for vendor in vendors:
                    if vendor.strip():
                        vendor_reference[vendor.strip().lower()] = col
                        vendor_reference[vendor.strip()] = col
        
        print(f"✅ Loaded {len(vendor_reference)} vendor references")
        return vendor_reference
    except Exception as e:
        print(f"❌ Error loading vendor reference: {e}")
        return {}

def is_invoice_email(subject, body):
    """Check if email is an invoice based on subject and body"""
    text = f"{subject} {body}".lower()
    return any(keyword in text for keyword in INVOICE_KEYWORDS)

def decode_email_subject(subject):
    """Decode email subject if it's encoded"""
    try:
        decoded = email.header.decode_header(subject)
        decoded_subject = ""
        for part, encoding in decoded:
            if isinstance(part, bytes):
                if encoding:
                    decoded_subject += part.decode(encoding)
                else:
                    decoded_subject += part.decode('utf-8', errors='ignore')
            else:
                decoded_subject += part
        return decoded_subject
    except:
        return subject

def extract_text_from_pdf(pdf_data):
    """Extract text content from PDF data"""
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_data))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"   ⚠️  Error extracting PDF text: {e}")
        return ""

def extract_attachment_info(email_message):
    """Extract attachment information and content from email"""
    attachments = []
    attachment_texts = []
    
    if email_message.is_multipart():
        for part in email_message.walk():
            if part.get_content_disposition() == 'attachment':
                filename = part.get_filename()
                content_type = part.get_content_type()
                payload = part.get_payload(decode=True)
                size = len(payload) if payload else 0
                
                attachment_info = {
                    'filename': filename or 'Unknown',
                    'content_type': content_type,
                    'size_bytes': size,
                    'payload': payload  # Include payload for AWS Bedrock Vision
                }
                
                # Extract text from PDF attachments
                if content_type == 'application/pdf' and payload:
                    print(f"   📄 Processing PDF: {filename}")
                    pdf_text = extract_text_from_pdf(payload)
                    if pdf_text:
                        attachment_info['extracted_text'] = pdf_text[:1000] + "..." if len(pdf_text) > 1000 else pdf_text
                        attachment_texts.append(pdf_text)
                        print(f"   ✅ Extracted {len(pdf_text)} characters from PDF")
                
                attachments.append(attachment_info)
    
    return attachments if attachments else ['None'], attachment_texts

def extract_vendor_name_improved(subject, from_email, body, pdf_analysis=None, pdf_data=None, openai_extractor=None):
    """Extract vendor name with improved logic"""
    # First, try to get hotel name from PDF analysis if available
    hotel_name = None
    if pdf_analysis:
        for pdf_key, pdf_data in pdf_analysis.items():
            if pdf_data.get('hotel_name') and pdf_data['hotel_name'] != 'Not Found':
                hotel_name = pdf_data['hotel_name']
                break
    
    # If no hotel name from PDF, try AWS Bedrock Vision
    if not hotel_name and pdf_data and openai_extractor and openai_extractor.enabled:
        try:
            hotel_name = openai_extractor.extract_property_name_from_pdf(pdf_data)
            if hotel_name and hotel_name != 'Not Found':
                print(f"   🤖 AWS Bedrock Vision extracted hotel: {hotel_name}")
        except Exception as e:
            print(f"   ⚠️  AWS Bedrock Vision error: {e}")
    
    # If we have a hotel name, use it
    if hotel_name and hotel_name != 'Not Found':
        return hotel_name
    
    # Fallback to email-based extraction
    text_to_search = f"{subject} {from_email} {body}".lower()
    
    # Look for common hotel/property indicators
    hotel_indicators = [
        'hotel', 'resort', 'inn', 'chalet', 'executive', 'apartments',
        'suites', 'palace', 'tower', 'plaza', 'central', 'grand',
        'premier', 'luxury', 'boutique', 'international', 'airport',
        'city', 'garden', 'park', 'view', 'heights', 'manor', 'villa',
        'house', 'lodge', 'court', 'square', 'mall', 'center', 'centre',
        'complex', 'building', 'towers'
    ]
    
    for indicator in hotel_indicators:
        if indicator in text_to_search:
            # Extract the word before the indicator
            pattern = r'(\w+)\s+' + indicator
            match = re.search(pattern, text_to_search)
            if match:
                potential_name = match.group(1).title() + ' ' + indicator.title()
                return potential_name
    
    # If no hotel name found, return generic
    return 'Unknown Vendor'

def extract_comprehensive_email_data(email_message, subject, from_email, body, openai_extractor=None):
    """Extract comprehensive information from email and attachments"""
    # Extract email headers
    to_email = email_message.get('To', '')
    cc_email = email_message.get('Cc', '')
    bcc_email = email_message.get('Bcc', '')
    reply_to = email_message.get('Reply-To', '')
    message_id = email_message.get('Message-ID', '')
    in_reply_to = email_message.get('In-Reply-To', '')
    references = email_message.get('References', '')
    
    # Extract attachments
    attachments, attachment_texts = extract_attachment_info(email_message)
    
    # Process PDF attachments with AWS Bedrock Vision
    pdf_analysis = {}
    if attachments and attachments != ['None']:
        for i, attachment in enumerate(attachments):
            if attachment.get('content_type') == 'application/pdf' and attachment.get('payload'):
                print(f"   🔍 Analyzing PDF {i+1} with AWS Bedrock Vision...")
                
                # Extract text for fallback
                pdf_text = extract_text_from_pdf(attachment['payload'])
                
                # Try AWS Bedrock Vision first
                openai_data = None
                if openai_extractor and openai_extractor.enabled:
                    try:
                        openai_data = openai_extractor.extract_comprehensive_invoice_data_from_pdf(attachment['payload'])
                        if openai_data:
                            print(f"   ✅ AWS Bedrock Vision extracted data for PDF {i+1}")
                    except Exception as e:
                        print(f"   ⚠️  AWS Bedrock Vision error for PDF {i+1}: {e}")
                
                # Create PDF data structure
                pdf_data = {
                    'hotel_name': 'Not Found',
                    'guest_name': 'Not Found',
                    'bill_number': 'Not Found',
                    'bill_date': 'Not Found',
                    'arrival_date': 'Not Found',
                    'departure_date': 'Not Found',
                    'room_number': 'Not Found',
                    'number_of_pax': 'Not Found',
                    'total_amount': 'Not Found',
                    'gst_number': 'Not Found',
                    'pan_number': 'Not Found',
                    'hotel_address': 'Not Found',
                    'hotel_phone': 'Not Found',
                    'hotel_email': 'Not Found',
                    'filename': attachment.get('filename', f'pdf_{i+1}'),
                    'payload': attachment.get('payload')  # Keep PDF data for upload
                }
                
                # Use AWS Bedrock data if available
                if openai_data:
                    pdf_data['openai_extracted_data'] = openai_data
                    # Parse AWS Bedrock response to extract structured data
                    if 'HOTEL:' in openai_data:
                        hotel_match = re.search(r'HOTEL:\s*([^\n]+)', openai_data)
                        if hotel_match:
                            pdf_data['hotel_name'] = hotel_match.group(1).strip()
                    
                    if 'GUEST:' in openai_data:
                        guest_match = re.search(r'GUEST:\s*([^\n]+)', openai_data)
                        if guest_match:
                            pdf_data['guest_name'] = guest_match.group(1).strip()
                    
                    if 'BILL NO:' in openai_data:
                        bill_match = re.search(r'BILL NO:\s*([^\n]+)', openai_data)
                        if bill_match:
                            pdf_data['bill_number'] = bill_match.group(1).strip()
                    
                    if 'AMOUNT:' in openai_data:
                        amount_match = re.search(r'AMOUNT:\s*([^\n]+)', openai_data)
                        if amount_match:
                            pdf_data['total_amount'] = amount_match.group(1).strip()
                    
                    if 'ROOM:' in openai_data:
                        room_match = re.search(r'ROOM:\s*([^\n]+)', openai_data)
                        if room_match:
                            pdf_data['room_number'] = room_match.group(1).strip()
                    
                    if 'GUESTS:' in openai_data:
                        guests_match = re.search(r'GUESTS:\s*([^\n]+)', openai_data)
                        if guests_match:
                            pdf_data['number_of_pax'] = guests_match.group(1).strip()
                    
                    if 'CHECK-IN:' in openai_data:
                        checkin_match = re.search(r'CHECK-IN:\s*([^\n]+)', openai_data)
                        if checkin_match:
                            pdf_data['arrival_date'] = checkin_match.group(1).strip()
                    
                    if 'CHECK-OUT:' in openai_data:
                        checkout_match = re.search(r'CHECK-OUT:\s*([^\n]+)', openai_data)
                        if checkout_match:
                            pdf_data['departure_date'] = checkout_match.group(1).strip()
                    
                    if 'BILL DATE:' in openai_data:
                        billdate_match = re.search(r'BILL DATE:\s*([^\n]+)', openai_data)
                        if billdate_match:
                            pdf_data['bill_date'] = billdate_match.group(1).strip()
                    
                    if 'GST:' in openai_data:
                        gst_match = re.search(r'GST:\s*([^\n]+)', openai_data)
                        if gst_match:
                            pdf_data['gst_number'] = gst_match.group(1).strip()
                else:
                    print(f"   ❌ PDF couldn't be processed")
                    pdf_data['hotel_name'] = 'PDF couldn\'t be processed'
                
                pdf_analysis[f'pdf_{i+1}'] = pdf_data
    
    # Basic text extraction for email metadata only
    combined_text = body + ' ' + subject
    amounts = extract_amounts(combined_text)
    urgency_level = extract_urgency_level(combined_text)
    
    return {
        'to_email': to_email,
        'cc_email': cc_email,
        'bcc_email': bcc_email,
        'reply_to': reply_to,
        'message_id': message_id,
        'in_reply_to': in_reply_to,
        'references': references,
        'attachments': attachments,
        'attachment_texts': attachment_texts,
        'pdf_analysis': pdf_analysis,
        'amounts': amounts,
        'urgency_level': urgency_level
    }

def extract_amounts(text):
    """Extract comprehensive monetary amounts from text"""
    amount_patterns = [
        # Currency symbols with amounts
        r'[\$₹€£]\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*[\$₹€£]',
        
        # Indian Rupee patterns
        r'₹\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'Rs\.?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'INR\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        
        # Generic number patterns that could be amounts
        r'(?<!\d)(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)(?!\d)',
    ]
    
    amounts = []
    for pattern in amount_patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                # Clean and convert to float
                clean_amount = match.replace(',', '')
                amount = float(clean_amount)
                # Filter out very small amounts that are likely not real prices
                if amount >= 1.0:
                    amounts.append(amount)
            except ValueError:
                continue
    
    return list(set(amounts))  # Remove duplicates

def extract_urgency_level(text):
    """Extract urgency level from text"""
    urgency_keywords = {
        'URGENT': ['urgent', 'asap', 'immediately', 'rush', 'priority'],
        'HIGH': ['high priority', 'important', 'critical'],
        'NORMAL': ['normal', 'standard', 'regular']
    }
    
    text_lower = text.lower()
    for level, keywords in urgency_keywords.items():
        if any(keyword in text_lower for keyword in keywords):
            return level
    
    return 'NORMAL'

def find_vendor_column(vendor_name, vendor_reference):
    """Find the appropriate column for a vendor using fuzzy matching"""
    if not vendor_name or vendor_name == 'Unknown Vendor':
        return 'UNASSIGNED'
    
    print(f"   🔍 Looking for vendor: '{vendor_name}'")
    
    # Direct match first (exact match)
    if vendor_name.lower() in vendor_reference:
        result = vendor_reference[vendor_name.lower()]
        print(f"   ✅ Direct match found: {result}")
        return result
    
    # Try partial matching for hotel names
    vendor_lower = vendor_name.lower()
    best_partial_match = None
    best_partial_score = 0
    
    for ref_name, column in vendor_reference.items():
        ref_lower = ref_name.lower()
        
        # Check if any significant word from the hotel name matches
        hotel_words = vendor_lower.split()
        ref_words = ref_lower.split()
        
        # Check for significant word matches (words longer than 3 characters)
        significant_matches = 0
        for hotel_word in hotel_words:
            if len(hotel_word) > 3:  # Only consider significant words
                for ref_word in ref_words:
                    if len(ref_word) > 3:
                        # Check if words are similar (fuzzy match)
                        if fuzz.ratio(hotel_word, ref_word) > 80:
                            significant_matches += 1
                            break
        
        # Also check for partial string matches (for cases like "Sapphire" in "SSAPPHIRE")
        partial_match_score = 0
        for hotel_word in hotel_words:
            if len(hotel_word) > 4:  # Only consider longer words
                for ref_word in ref_words:
                    if len(ref_word) > 4:
                        # Check if one word contains the other or vice versa
                        if hotel_word in ref_word or ref_word in hotel_word:
                            partial_match_score += 1
                        # Also check fuzzy match for partial words
                        elif fuzz.ratio(hotel_word, ref_word) > 70:
                            partial_match_score += 0.5
        
        # Calculate total score
        total_score = significant_matches + partial_match_score
        
        if total_score > best_partial_score:
            best_partial_score = total_score
            best_partial_match = (ref_name, column)
        
        # If we have at least 2 significant word matches, consider it a match
        if significant_matches >= 2:
            print(f"   ✅ Partial match found: '{ref_name}' -> {column} ({significant_matches} words match)")
            return column
    
    # If we have a good partial match, use it
    if best_partial_match and best_partial_score >= 1.5:
        ref_name, column = best_partial_match
        print(f"   ✅ Best partial match found: '{ref_name}' -> {column} (score: {best_partial_score})")
        return column
    
    # Fuzzy matching on the full name
    best_match = process.extractOne(vendor_name, list(vendor_reference.keys()), scorer=fuzz.ratio)
    if best_match and best_match[1] >= 60:  # Lowered threshold to 60%
        print(f"   ✅ Fuzzy match found: '{best_match[0]}' -> {vendor_reference[best_match[0]]} ({best_match[1]}% match)")
        return vendor_reference[best_match[0]]
    
    print(f"   ❌ No match found for: '{vendor_name}'")
    return 'UNASSIGNED'

def get_drive_folder_name(excel_column):
    """Convert Excel column name to Google Drive folder name"""
    # Now MOKSHITHA maps directly to MOKSHITHA folder
    return excel_column

def upload_pdf_to_drive(pdf_data, assigned_column, drive_uploader):
    """Upload PDF file to Google Drive folder"""
    if not drive_uploader or not pdf_data.get('payload'):
        return None
    
    try:
        # Create safe filename with date prefix
        original_filename = pdf_data.get('filename', 'unknown.pdf')
        safe_filename = re.sub(r'[^\w\s\-\.]', '', original_filename).strip()
        safe_filename = safe_filename.replace(' ', '_')
        
        # Add date prefix in YYYYMMDD format
        processing_date = datetime.now().strftime('%Y%m%d')
        date_filename = f"{processing_date}_{safe_filename}"
        
        # Upload to Google Drive - convert Excel column to Drive folder name
        folder_name = get_drive_folder_name(assigned_column) if assigned_column != 'Not Found' else 'UNASSIGNED'
        print(f"   🔧 DEBUG: assigned_column='{assigned_column}', folder_name='{folder_name}'")
        print(f"   📅 Original: {original_filename} -> Date format: {date_filename}")
        file_id = drive_uploader.upload_pdf_data(pdf_data['payload'], date_filename, folder_name)
        
        if file_id:
            print(f"   ☁️  Uploaded PDF to Google Drive: {folder_name}/{date_filename}")
            return file_id
        else:
            print(f"   ⚠️  Failed to upload PDF to Google Drive")
            return None
            
    except Exception as e:
        print(f"   ❌ Error uploading PDF to Drive: {e}")
        return None

def update_excel_sheet(invoice_data_list, drive_uploader):
    """Update the master Excel sheet with new invoice data"""
    if not invoice_data_list:
        print("⚠️  No invoice data to save")
        return None
    
    # Download the current master Excel file
    master_file_id = '1Nb-K8ROUun1qI4QajNMTTlk6Pi03AlbT'
    try:
        # Download current file
        file_content = drive_uploader.service.files().get_media(fileId=master_file_id).execute()
        with open('temp_master.xlsx', 'wb') as f:
            f.write(file_content)
        print("📊 Downloaded current master Excel file")
        
        # Load existing data
        df = pd.read_excel('temp_master.xlsx')
        print(f"📊 Current Excel has {len(df)} rows")
        
        # Prepare new data to append
        new_rows = []
        for invoice in invoice_data_list:
            for pdf_key, pdf_data in invoice.get('pdf_analysis', {}).items():
                # Skip if PDF couldn't be processed
                if pdf_data.get('hotel_name') == 'PDF couldn\'t be processed':
                    continue
                
                # Create new row data with proper date formatting
                processing_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                email_from = invoice.get('from_email', '').replace('<', '').replace('>', '')
                email_subject = invoice.get('subject', '')[:50] + '...' if len(invoice.get('subject', '')) > 50 else invoice.get('subject', '')
                assigned_folder = invoice.get('assigned_column', 'UNASSIGNED')
                hotel_name = pdf_data.get('hotel_name', 'Not Found')
                guest_name = pdf_data.get('guest_name', 'Not Found')
                bill_number = pdf_data.get('bill_number', 'Not Found')
                
                # Format bill date properly
                bill_date = pdf_data.get('bill_date', 'Not Found')
                if bill_date != 'Not Found' and bill_date:
                    try:
                        # Try to parse and reformat the date
                        if '/' in str(bill_date):
                            # Handle DD/MM/YY or DD/MM/YYYY format
                            date_parts = str(bill_date).split('/')
                            if len(date_parts) == 3:
                                day, month, year = date_parts
                                if len(year) == 2:
                                    year = '20' + year
                                bill_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    except:
                        pass  # Keep original if parsing fails
                
                room_number = pdf_data.get('room_number', 'Not Found')
                number_of_guests = pdf_data.get('number_of_pax', 'Not Found')
                
                # Format check-in and check-out dates properly
                check_in_date = pdf_data.get('arrival_date', 'Not Found')
                if check_in_date != 'Not Found' and check_in_date:
                    try:
                        if '/' in str(check_in_date):
                            date_parts = str(check_in_date).split('/')
                            if len(date_parts) == 3:
                                day, month, year = date_parts
                                if len(year) == 2:
                                    year = '20' + year
                                check_in_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    except:
                        pass
                
                check_out_date = pdf_data.get('departure_date', 'Not Found')
                if check_out_date != 'Not Found' and check_out_date:
                    try:
                        if '/' in str(check_out_date):
                            date_parts = str(check_out_date).split('/')
                            if len(date_parts) == 3:
                                day, month, year = date_parts
                                if len(year) == 2:
                                    year = '20' + year
                                check_out_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    except:
                        pass
                total_amount = pdf_data.get('total_amount', 'Not Found')
                gst_number = pdf_data.get('gst_number', 'Not Found')
                pan_number = pdf_data.get('pan_number', 'Not Found')
                # Use date-formatted filename for Excel
                original_filename = pdf_data.get('filename', '')
                processing_date_for_filename = datetime.now().strftime('%Y%m%d')
                safe_filename = re.sub(r'[^\w\s\-\.]', '', original_filename).strip().replace(' ', '_')
                pdf_filename = f"{processing_date_for_filename}_{safe_filename}"
                drive_file_id = pdf_data.get('drive_file_id', '')
                urgency_level = invoice.get('urgency_level', 'NORMAL')
                amounts_found = ', '.join(map(str, invoice.get('amounts', [])))
                
                # Get next S.No
                next_s_no = len(df) + len(new_rows) + 1
                
                new_row = {
                    'S.No': next_s_no,
                    'Processing Date': processing_date,
                    'Email From': email_from,
                    'Email Subject': email_subject,
                    'Assigned To': assigned_folder,
                    'Hotel Name': hotel_name,
                    'Guest Name': guest_name,
                    'Bill Number': bill_number,
                    'Bill Date': bill_date,
                    'Room Number': room_number,
                    'Number of Guests': number_of_guests,
                    'Check-in Date': check_in_date,
                    'Check-out Date': check_out_date,
                    'Total Amount': total_amount,
                    'GST Number': gst_number,
                    'PAN Number': pan_number,
                    'PDF Filename': pdf_filename,
                    'Drive File ID': drive_file_id,
                    'Urgency Level': urgency_level,
                    'Amounts Found': amounts_found
                }
                new_rows.append(new_row)
        
        if not new_rows:
            print("⚠️  No valid invoice data to add to Excel")
            return None
        
        # Append new data to existing DataFrame
        new_df = pd.DataFrame(new_rows)
        updated_df = pd.concat([df, new_df], ignore_index=True)
        
        # Save updated file with improved formatting
        with pd.ExcelWriter('temp_master_updated.xlsx', engine='openpyxl') as writer:
            updated_df.to_excel(writer, sheet_name='Invoice Data', index=False)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Invoice Data']
            
            # Set column widths for better readability
            column_widths = {
                'A': 8,   # S.No
                'B': 20,  # Processing Date
                'C': 25,  # Email From
                'D': 40,  # Email Subject
                'E': 15,  # Assigned To
                'F': 35,  # Hotel Name
                'G': 25,  # Guest Name
                'H': 18,  # Bill Number
                'I': 15,  # Bill Date
                'J': 12,  # Room Number
                'K': 15,  # Number of Guests
                'L': 15,  # Check-in Date
                'M': 15,  # Check-out Date
                'N': 18,  # Total Amount
                'O': 20,  # GST Number
                'P': 15,  # PAN Number
                'Q': 25,  # PDF Filename
                'R': 25,  # Drive File ID
                'S': 12,  # Urgency Level
                'T': 30   # Amounts Found
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Add professional formatting
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Apply header formatting
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply borders to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Data rows
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Freeze the header row
            worksheet.freeze_panes = "A2"
            
            # Auto-fit row heights
            for row in worksheet.iter_rows():
                max_height = 0
                for cell in row:
                    if cell.value:
                        # Calculate approximate height based on content
                        lines = str(cell.value).count('\n') + 1
                        height = max(15, lines * 15)  # Minimum 15, 15 per line
                        max_height = max(max_height, height)
                if max_height > 0:
                    worksheet.row_dimensions[row[0].row].height = min(max_height, 60)  # Max 60
            
            # Format date columns properly
            from openpyxl.styles import NamedStyle
            date_style = NamedStyle(name="date_style")
            date_style.number_format = 'YYYY-MM-DD'
            
            # Apply date formatting to date columns (B, I, L, M)
            date_columns = ['B', 'I', 'L', 'M']  # Processing Date, Bill Date, Check-in, Check-out
            for col in date_columns:
                for row in range(2, worksheet.max_row + 1):  # Skip header row
                    cell = worksheet[f'{col}{row}']
                    if cell.value and str(cell.value) != 'Not Found':
                        try:
                            # Try to format as date if it looks like a date
                            cell_value = str(cell.value)
                            if '-' in cell_value or '/' in cell_value:
                                cell.number_format = 'YYYY-MM-DD'
                        except:
                            pass
        
        print(f"📊 Updated Excel with {len(new_rows)} new rows")
        
        # Upload back to Google Drive
        from googleapiclient.http import MediaFileUpload
        media_body = MediaFileUpload('temp_master_updated.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        updated_file = drive_uploader.service.files().update(
            fileId=master_file_id,
            media_body=media_body
        ).execute()
        
        print(f"📊 Successfully updated master Excel file in Google Drive")
        print(f"🔗 Link: https://drive.google.com/file/d/{master_file_id}/view")
        
        # Clean up temp files
        os.remove('temp_master.xlsx')
        os.remove('temp_master_updated.xlsx')
        
        return f"https://drive.google.com/file/d/{master_file_id}/view"
        
    except Exception as e:
        print(f"❌ Error updating Excel file: {e}")
        return None

def process_invoices():
    """Main function to process invoices from Gmail"""
    print("📧 FINAL INVOICE PROCESSOR - COMPLETE FLOW")
    print("=" * 50)
    print("1. 📧 Check unread emails")
    print("2. 🤖 Process invoices with AWS Bedrock Vision")
    print("3. ☁️  Upload PDFs to Google Drive folders")
    print("4. 📊 Update Excel sheet with extracted data")
    print("=" * 50)
    
    # Initialize AWS Bedrock extractor
    openai_extractor = OpenAIPropertyExtractor()
    if openai_extractor.enabled:
        print("🤖 AWS Bedrock Vision enabled for PDF analysis")
    else:
        print("⚠️  AWS Bedrock Vision disabled - using text analysis only")
    
    # Initialize Google Drive uploader
    drive_uploader = None
    if os.getenv('ENABLE_GOOGLE_DRIVE_UPLOAD', 'false').lower() == 'true':
        try:
            drive_uploader = GoogleDriveUploader()
            if drive_uploader.authenticate():
                drive_uploader.setup_folders()
                print("✅ Google Drive initialized successfully")
            else:
                print("❌ Google Drive authentication failed")
                return
        except Exception as e:
            print(f"❌ Google Drive setup failed: {e}")
            return
    else:
        print("❌ Google Drive upload disabled")
        return
    
    # Load vendor reference data
    vendor_reference = load_vendor_reference()
    print(f"✅ Loaded {len(vendor_reference)} vendor references")
    
    # Connect to Gmail
    try:
        mail = imaplib.IMAP4_SSL(os.getenv('IMAP_SERVER'), int(os.getenv('IMAP_PORT')))
        mail.login(os.getenv('GMAIL_EMAIL'), os.getenv('GMAIL_PASSWORD'))
        mail.select('inbox')
        print("✅ Connected to Gmail successfully!")
    except Exception as e:
        print(f"❌ Gmail connection failed: {e}")
        return
    
    # Search for unread emails
    try:
        status, messages = mail.search(None, 'UNSEEN')
        unread_emails = messages[0].split()
        print(f"📧 Unread emails: {len(unread_emails)}")
        
        # Get emails from last N days - use a more robust approach
        days_to_search = int(os.getenv('DAYS_TO_SEARCH', '7'))
        
        # Use a fixed reference date to avoid system date issues
        current_date = datetime.now()
        if current_date.year > 2024:  # If system date seems wrong, use a fixed date
            reference_date = datetime(2024, 10, 6)  # Use a known good date
            print(f"⚠️  System date appears incorrect ({current_date.strftime('%Y-%m-%d')}), using reference date")
        else:
            reference_date = current_date
            
        since_date = (reference_date - timedelta(days=days_to_search)).strftime('%d-%b-%Y')
        print(f"🔍 Searching for unread emails since: {since_date}")
        
        # Try different date formats if the first one doesn't work
        try:
            status, messages = mail.search(None, f'UNSEEN SINCE {since_date}')
            recent_emails = messages[0].split()
            print(f"📧 Found {len(recent_emails)} unread emails since {since_date}")
        except Exception as e:
            print(f"⚠️  First date format failed: {e}")
            # Try alternative date format
            since_date_alt = (reference_date - timedelta(days=days_to_search)).strftime('%d-%b-%Y')
            try:
                status, messages = mail.search(None, f'UNSEEN SINCE {since_date_alt}')
                recent_emails = messages[0].split()
                print(f"📧 Found {len(recent_emails)} unread emails since {since_date_alt}")
            except Exception as e2:
                print(f"⚠️  Alternative date format also failed: {e2}")
                # Fallback to just UNSEEN without date filter
                status, messages = mail.search(None, 'UNSEEN')
                recent_emails = messages[0].split()
                print(f"📧 Fallback: Found {len(recent_emails)} unread emails (no date filter)")
        
    except Exception as e:
        print(f"❌ Error searching emails: {e}")
        mail.close()
        mail.logout()
        return
    
    # Process emails
    processed_invoice_ids = []
    invoice_count = 0
    folder_counts = {'KUMAR': 0, 'LAKSHMI': 0, 'MOKSHITHA': 0, 'SANDHYA': 0, 'UNASSIGNED': 0}
    all_invoice_data = []  # Store all invoice data for Excel
    
    for i, email_id in enumerate(recent_emails):
        try:
            print(f"\n📨 Processing email {i+1}/{len(recent_emails)}")
            
            # Fetch email
            status, msg_data = mail.fetch(email_id, '(RFC822)')
            email_message = email.message_from_bytes(msg_data[0][1])
            
            # Extract basic email info
            subject = decode_email_subject(email_message.get('Subject', ''))
            from_email = email_message.get('From', '')
            body = ""
            
            if email_message.is_multipart():
                for part in email_message.walk():
                    if part.get_content_type() == "text/plain":
                        body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                        break
            else:
                body = email_message.get_payload(decode=True).decode('utf-8', errors='ignore')
            
            print(f"   From: {from_email}")
            print(f"   Subject: {subject}")
            
            # Check if it's an invoice email
            if is_invoice_email(subject, body):
                print("   🎯 INVOICE DETECTED!")
                invoice_count += 1
                
                # Extract comprehensive email data
                comprehensive_data = extract_comprehensive_email_data(email_message, subject, from_email, body, openai_extractor)
                
                # Extract vendor name using improved method
                pdf_data_for_vision = None
                if comprehensive_data.get('attachments'):
                    for attachment in comprehensive_data['attachments']:
                        if attachment.get('content_type') == 'application/pdf' and attachment.get('payload'):
                            pdf_data_for_vision = attachment['payload']
                            break
                
                vendor_name = extract_vendor_name_improved(
                    subject, from_email, body, 
                    comprehensive_data.get('pdf_analysis'), 
                    pdf_data_for_vision, 
                    openai_extractor
                )
                
                # Find the correct column for this vendor
                assigned_column = find_vendor_column(vendor_name, vendor_reference)
                
                print(f"   🏨 Extracted vendor: {vendor_name}")
                print(f"   📁 Assigned to: {assigned_column}")
                print(f"   🔧 DEBUG: assigned_column type: {type(assigned_column)}, value: '{assigned_column}'")
                
                # Extract amounts for display
                amounts = comprehensive_data.get('amounts', [])
                if amounts:
                    print(f"   💰 Amounts found: {amounts}")
                
                # Extract urgency level
                urgency = comprehensive_data.get('urgency_level', 'NORMAL')
                print(f"   ⚡ Urgency: {urgency}")
                
                # Process each PDF invoice
                if 'pdf_analysis' in comprehensive_data and comprehensive_data['pdf_analysis']:
                    pdf_count = 0
                    for pdf_key, pdf_data in comprehensive_data['pdf_analysis'].items():
                        # Skip if PDF couldn't be processed
                        if pdf_data.get('hotel_name') == 'PDF couldn\'t be processed':
                            print(f"   ⚠️  Skipping PDF {pdf_count + 1} - couldn't be processed")
                            continue
                            
                        pdf_count += 1
                        
                        # Upload PDF to Google Drive
                        if drive_uploader:
                            file_id = upload_pdf_to_drive(pdf_data, assigned_column, drive_uploader)
                            if file_id:
                                pdf_data['drive_file_id'] = file_id
                                folder_counts[assigned_column if assigned_column != 'Not Found' else 'UNASSIGNED'] += 1
                                print(f"   💾 Uploaded PDF {pdf_count}: {pdf_data.get('filename', 'unknown.pdf')}")
                            else:
                                print(f"   ❌ Failed to upload PDF {pdf_count}")
                        else:
                            print(f"   ⚠️  Google Drive not available - PDF not uploaded")
                    
                    if pdf_count > 0:
                        print(f"   ✅ Processed {pdf_count} PDF(s) for: {assigned_column} folder")
                else:
                    print("   ⚠️  No PDF analysis data found")
                
                # Store invoice data for Excel
                invoice_data = {
                    'from_email': from_email,
                    'subject': subject,
                    'assigned_column': assigned_column,
                    'urgency_level': urgency,
                    'amounts': amounts,
                    'pdf_analysis': comprehensive_data.get('pdf_analysis', {})
                }
                all_invoice_data.append(invoice_data)
                
                # Only mark invoice emails as read
                processed_invoice_ids.append(email_id)
            else:
                print("   📄 Regular email (not an invoice) - leaving unread")
            
        except Exception as e:
            print(f"   ❌ Error processing email {i+1}: {e}")
            continue
    
    # Mark processed emails as read
    if processed_invoice_ids:
        try:
            print(f"\n📧 Marking {len(processed_invoice_ids)} invoice emails as read...")
            for email_id in processed_invoice_ids:
                mail.store(email_id, '+FLAGS', '\\Seen')
            print("✅ Invoice emails marked as read")
        except Exception as e:
            print(f"⚠️  Error marking emails as read: {e}")
    
    # Update Excel sheet with all invoice data
    if all_invoice_data:
        excel_link = update_excel_sheet(all_invoice_data, drive_uploader)
        if excel_link:
            print(f"📊 Successfully updated Excel sheet: {excel_link}")
        else:
            print("❌ Failed to update Excel sheet")
    else:
        print("⚠️  No invoice data to update in Excel")
    
    # Final statistics
    try:
        status, messages = mail.search(None, 'UNSEEN')
        unread_after = len(messages[0].split()) if messages[0] else 0
        print(f"\n📧 Unread emails after processing: {unread_after}")
        print(f"📊 Invoice emails processed: {len(processed_invoice_ids)}")
        print(f"🎯 Invoices found: {invoice_count}")
        
        # Print summary
        print(f"\n📊 INVOICE SUMMARY:")
        print("-" * 30)
        for folder, count in folder_counts.items():
            if count > 0:
                print(f"📁 {folder}: {count} PDFs uploaded")
        
    except Exception as e:
        print(f"⚠️  Error getting final statistics: {e}")
    
    mail.close()
    mail.logout()

if __name__ == "__main__":
    process_invoices()
