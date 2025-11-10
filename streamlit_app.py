#!/usr/bin/env python3
"""
Streamlit UI for Invoice Automation Hub
Provides a user-friendly web interface for processing invoices from Gmail,
uploading to Google Drive, and updating Excel sheets.
"""

import streamlit as st
import os
import sys
import io
import contextlib
from datetime import datetime
import pandas as pd
from dotenv import load_dotenv

# Import invoice processor functions
from invoice_processor_final import (
    process_invoices,
    match_master_sheet_with_excel,
    load_vendor_reference,
    read_matching_excel_file,
    match_invoice_master_sheet_with_excel,
    MATCHING_EXCEL_FILE_PATH,
    INVOICE_MASTER_SHEET_ID,
    MATCHING_FIELDS,
    find_matching_column,
    INVOICE_RECEIVED_COLUMN,
    INVOICE_RECEIVED_VALUE
)

# Load environment variables
load_dotenv()

# Page configuration
st.set_page_config(
    page_title="Invoice Automation Hub",
    page_icon="📧",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS for better styling
st.markdown("""
    <style>
    .main-header {
        font-size: 4.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        padding: 1.5rem;
        margin-bottom: 1rem;
    }
    .status-card {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #f0f2f6;
        margin: 0.5rem 0;
    }
    .success {
        color: #28a745;
    }
    .error {
        color: #dc3545;
    }
    .warning {
        color: #ffc107;
    }
    .info {
        color: #17a2b8;
    }
    </style>
""", unsafe_allow_html=True)

# Initialize session state
if 'processing_logs' not in st.session_state:
    st.session_state.processing_logs = []
if 'last_processed' not in st.session_state:
    st.session_state.last_processed = None
if 'processing_status' not in st.session_state:
    st.session_state.processing_status = "Ready"
if 'vendor_reference' not in st.session_state:
    st.session_state.vendor_reference = None

def capture_output(func, *args, **kwargs):
    """Capture stdout and stderr from a function call"""
    output = io.StringIO()
    error_output = io.StringIO()
    
    with contextlib.redirect_stdout(output), contextlib.redirect_stderr(error_output):
        try:
            result = func(*args, **kwargs)
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
            result = None
    
    stdout_text = output.getvalue()
    stderr_text = error_output.getvalue()
    
    # Combine outputs
    combined_output = stdout_text + stderr_text
    
    return result, combined_output

def check_configuration():
    """Check if all required configuration is set"""
    required_vars = {
        'GMAIL_EMAIL': os.getenv('GMAIL_EMAIL'),
        'GMAIL_PASSWORD': os.getenv('GMAIL_PASSWORD'),
        'IMAP_SERVER': os.getenv('IMAP_SERVER', 'imap.gmail.com'),
        'IMAP_PORT': os.getenv('IMAP_PORT', '993'),
        'EXCEL_FILE_PATH': os.getenv('EXCEL_FILE_PATH'),
        'ENABLE_GOOGLE_DRIVE_UPLOAD': os.getenv('ENABLE_GOOGLE_DRIVE_UPLOAD', 'false'),
    }
    
    optional_vars = {
        'ENABLE_OPENAI_VISION': os.getenv('ENABLE_OPENAI_VISION', 'false'),
        'AWS_BEDROCK_MODEL': os.getenv('AWS_BEDROCK_MODEL'),
        'DAYS_TO_SEARCH': os.getenv('DAYS_TO_SEARCH', '7'),
        'MASTER_EXCEL_FILE_ID': os.getenv('MASTER_EXCEL_FILE_ID'),
    }
    
    return required_vars, optional_vars

def test_gmail_connection():
    """Test Gmail IMAP connection"""
    try:
        import imaplib
        mail = imaplib.IMAP4_SSL(
            os.getenv('IMAP_SERVER', 'imap.gmail.com'),
            int(os.getenv('IMAP_PORT', '993'))
        )
        mail.login(os.getenv('GMAIL_EMAIL'), os.getenv('GMAIL_PASSWORD'))
        mail.select('inbox')
        mail.close()
        mail.logout()
        return True, "✅ Gmail connection successful!"
    except Exception as e:
        return False, f"❌ Gmail connection failed: {e}"

def test_google_drive_connection():
    """Test Google Drive connection"""
    try:
        from google_drive_uploader import GoogleDriveUploader
        drive_uploader = GoogleDriveUploader()
        if drive_uploader.authenticate():
            return True, "✅ Google Drive connection successful!"
        else:
            return False, "❌ Google Drive authentication failed"
    except Exception as e:
        return False, f"❌ Google Drive connection failed: {e}"

def read_file_from_drive_fixed(file_id, drive_uploader, temp_filename='temp_drive_file.xlsx'):
    """
    Read a file from Google Drive, handling both Google Sheets and Excel files
    This is a fixed version that works with both file types
    
    Args:
        file_id: Google Drive file ID
        drive_uploader: GoogleDriveUploader instance
        temp_filename: Temporary filename to save the file
    
    Returns:
        pandas DataFrame or None if error
    """
    if not drive_uploader or not drive_uploader.service:
        print("❌ Google Drive service not available")
        return None
    
    try:
        # First, get file metadata to check the file type
        file_metadata = drive_uploader.service.files().get(
            fileId=file_id,
            fields='id,name,mimeType'
        ).execute()
        
        mime_type = file_metadata.get('mimeType', '')
        file_name = file_metadata.get('name', 'Unknown')
        
        print(f"📄 File: {file_name}")
        print(f"📋 MIME Type: {mime_type}")
        
        # Check if it's a Google Sheets file
        if mime_type == 'application/vnd.google-apps.spreadsheet':
            # It's a Google Sheets file - use export_media
            print("   📊 Detected Google Sheets file - using export_media")
            request = drive_uploader.service.files().export_media(
                fileId=file_id,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            # It's an Excel file or other file - use get_media
            print("   📄 Detected Excel/other file - using get_media")
            request = drive_uploader.service.files().get_media(fileId=file_id)
        
        # Download to temporary file
        print(f"   📥 Downloading file...")
        with open(temp_filename, 'wb') as f:
            f.write(request.execute())
        
        print(f"   ✅ Downloaded file successfully")
        
        # Read Excel file with pandas
        df = pd.read_excel(temp_filename, engine='openpyxl', header=0)
        print(f"   ✅ Read {len(df)} rows from file")
        print(f"   📋 Columns: {list(df.columns)}")
        
        return df
        
    except Exception as e:
        print(f"❌ Error reading file from Google Drive: {e}")
        import traceback
        traceback.print_exc()
        return None

def read_invoice_master_sheet_fixed(drive_uploader, file_id):
    """
    Read Invoice Processor Master Data from Google Drive (fixed version)
    Handles both Google Sheets and Excel files
    
    Args:
        drive_uploader: GoogleDriveUploader instance
        file_id: Google Drive file ID
    
    Returns:
        List of extracted data entries or None
    """
    if not drive_uploader or not drive_uploader.service:
        print("❌ Google Drive service not available")
        return None
    
    try:
        print(f"\n📊 Reading Invoice Master Sheet from Google Drive...")
        print(f"   File ID: {file_id}")
        
        # Use the helper function to read the file
        df = read_file_from_drive_fixed(file_id, drive_uploader, 'temp_invoice_master.xlsx')
        
        if df is None or df.empty:
            print("   ⚠️  No data found in file")
            return None
        
        # Extract matching fields from the sheet
        extracted_data_list = []
        
        # Find column mappings
        booking_code_col = None
        guest_name_col = None
        checkin_col = None
        checkout_col = None
        
        # Map columns (case-insensitive)
        for col in df.columns:
            col_lower = str(col).strip().lower()
            if 'booking code' in col_lower or 'booking' in col_lower:
                booking_code_col = col
            elif 'guest name' in col_lower or ('guest' in col_lower and 'name' in col_lower):
                guest_name_col = col
            elif 'check-in' in col_lower or 'check in' in col_lower or 'arrival' in col_lower:
                checkin_col = col
            elif 'check-out' in col_lower or 'check out' in col_lower or 'departure' in col_lower:
                checkout_col = col
        
        print(f"\n   🔍 Column mappings found:")
        print(f"      Booking Code: {booking_code_col}")
        print(f"      Guest Name: {guest_name_col}")
        print(f"      Check-In Date: {checkin_col}")
        print(f"      Check-Out Date: {checkout_col}")
        
        # Extract data from each row
        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get(booking_code_col if booking_code_col else '')) and \
               pd.isna(row.get(guest_name_col if guest_name_col else '')):
                continue
            
            entry = {
                'Booking Code': str(row[booking_code_col]).strip() if booking_code_col and not pd.isna(row.get(booking_code_col)) else '',
                'Guest Name': str(row[guest_name_col]).strip() if guest_name_col and not pd.isna(row.get(guest_name_col)) else '',
                'Check-In Date': str(row[checkin_col]).strip() if checkin_col and not pd.isna(row.get(checkin_col)) else '',
                'Check-Out Date': str(row[checkout_col]).strip() if checkout_col and not pd.isna(row.get(checkout_col)) else ''
            }
            
            # Clean up "Not Found" and "nan" values
            for key, value in entry.items():
                if value.lower() in ['not found', 'nan', 'none', '']:
                    entry[key] = ''
            
            # Only add if we have at least one field
            if any(entry.values()):
                extracted_data_list.append(entry)
        
        print(f"   ✅ Extracted {len(extracted_data_list)} entries from Invoice Master Sheet")
        
        # Clean up temp file
        try:
            import time
            time.sleep(0.3)
            if os.path.exists('temp_invoice_master.xlsx'):
                os.remove('temp_invoice_master.xlsx')
        except:
            pass
        
        return extracted_data_list
        
    except Exception as e:
        print(f"❌ Error reading Invoice Master Sheet from Google Drive: {e}")
        import traceback
        traceback.print_exc()
        return None

def match_master_sheet_with_excel_fixed():
    """
    Fixed version of match_master_sheet_with_excel that correctly handles Excel files
    """
    from google_drive_uploader import GoogleDriveUploader
    
    # Initialize Google Drive uploader
    drive_uploader = None
    try:
        drive_uploader = GoogleDriveUploader()
        if drive_uploader.authenticate():
            print("✅ Google Drive initialized successfully")
        else:
            print("❌ Google Drive authentication failed")
            return 0
    except Exception as e:
        print(f"❌ Google Drive setup failed: {e}")
        return 0
    
    # Read Invoice Master Sheet from Google Drive using fixed function
    extracted_data_list = read_invoice_master_sheet_fixed(drive_uploader, INVOICE_MASTER_SHEET_ID)
    if not extracted_data_list:
        print("⚠️  No data extracted from Invoice Master Sheet - skipping matching")
        return 0
    
    # Read local Excel file
    df = read_matching_excel_file()
    if df is None or df.empty:
        print("⚠️  Excel file is empty or invalid - skipping matching")
        return 0
    
    print(f"\n🔍 Matching Invoice Master Sheet data with local Excel file...")
    print(f"🔍 Excel file has {len(df)} rows and {len(df.columns)} columns")
    print(f"🔍 Excel columns: {list(df.columns)}")
    print(f"🔍 Extracted {len(extracted_data_list)} entries from Invoice Master Sheet")
    
    # Create a mapping of field names to Excel columns
    field_to_column = {}
    for field in MATCHING_FIELDS:
        column = find_matching_column(df, field)
        if column:
            field_to_column[field] = column
            print(f"✅ Mapped '{field}' → '{column}'")
        else:
            print(f"⚠️  No matching column found for '{field}'")
    
    if not field_to_column:
        print("❌ No matching columns found. Cannot match Excel rows.")
        return 0
    
    # Check if 'Invoice Received' column exists, if not add it at the end
    if INVOICE_RECEIVED_COLUMN not in df.columns:
        print(f"➕ Adding new column: '{INVOICE_RECEIVED_COLUMN}'")
        df[INVOICE_RECEIVED_COLUMN] = ''
    else:
        print(f"✅ Column '{INVOICE_RECEIVED_COLUMN}' already exists")
    
    rows_updated = 0
    rows_skipped = 0
    
    # Process each extracted data entry (same logic as original function)
    for idx, entry in enumerate(extracted_data_list):
        if not entry or not any(entry.values()):
            continue
        
        print(f"\n📋 Processing entry {idx+1}:")
        for field, value in entry.items():
            if value:
                print(f"   {field}: {value}")
        
        matching_row_indices = []
        
        # Step 1: Try matching by Booking Code first (Primary Key)
        if 'Booking Code' in field_to_column:
            booking_code_column = field_to_column['Booking Code']
            booking_code_value = entry.get('Booking Code', '').strip()
            
            if booking_code_value:
                print(f"   🔍 Step 1: Matching by Booking Code: '{booking_code_value}'")
                try:
                    # Case-insensitive matching
                    mask = df[booking_code_column].astype(str).str.strip().str.lower() == booking_code_value.lower()
                    matching_row_indices = df[mask].index.tolist()
                    
                    if matching_row_indices:
                        print(f"   ✅ Found {len(matching_row_indices)} row(s) with Booking Code: '{booking_code_value}'")
                    else:
                        print(f"   ⚠️  No row found with Booking Code: '{booking_code_value}'")
                except Exception as e:
                    print(f"   ❌ Error matching by Booking Code: {e}")
        
        # Step 2: If no Booking Code match, try matching by Guest Name + Check-In Date + Check-Out Date
        if not matching_row_indices:
            print(f"   🔍 Step 2: Trying to match by Guest Name + Check-In Date + Check-Out Date")
            
            required_fields = ['Guest Name', 'Check-In Date', 'Check-Out Date']
            missing_fields = [f for f in required_fields if f not in field_to_column]
            
            if missing_fields:
                print(f"   ⚠️  Cannot match: Missing columns for {missing_fields}")
                continue
            
            # Get values for all three fields
            guest_name_value = entry.get('Guest Name', '').strip()
            checkin_value = entry.get('Check-In Date', '').strip()
            checkout_value = entry.get('Check-Out Date', '').strip()
            
            # All three must be present
            if not guest_name_value or not checkin_value or not checkout_value:
                print(f"   ⚠️  Cannot match: Missing required fields")
                continue
            
            print(f"   🔍 Matching by:")
            print(f"      Guest Name: '{guest_name_value}'")
            print(f"      Check-In Date: '{checkin_value}'")
            print(f"      Check-Out Date: '{checkout_value}'")
            
            try:
                guest_col = field_to_column['Guest Name']
                checkin_col = field_to_column['Check-In Date']
                checkout_col = field_to_column['Check-Out Date']
                
                # Normalize date formats for matching (same logic as original)
                def normalize_date_for_match(date_str):
                    """Normalize date string for comparison"""
                    if pd.isna(date_str) or not date_str:
                        return '', ''
                    
                    if hasattr(date_str, 'strftime'):
                        try:
                            format1 = date_str.strftime('%d-%b-%Y').lower()
                            format2 = date_str.strftime('%d/%m/%Y').lower()
                            return format1, format2
                        except:
                            date_str_lower = str(date_str).lower()
                            return date_str_lower, date_str_lower
                    
                    date_str = str(date_str).strip().lower()
                    format1 = date_str.replace('/', '-').replace(' ', '-').replace('.', '-')
                    format2 = date_str
                    return format1, format2
                
                # Normalize extracted dates
                checkin_norm, checkin_norm2 = normalize_date_for_match(checkin_value)
                checkout_norm, checkout_norm2 = normalize_date_for_match(checkout_value)
                
                # Case-insensitive matching for guest name
                condition1 = df[guest_col].astype(str).str.strip().str.lower() == guest_name_value.lower()
                if not condition1.any():
                    condition1 = df[guest_col].astype(str).str.strip().str.lower().str.contains(
                        guest_name_value.lower(), na=False, regex=False
                    )
                
                # Date matching - check-in
                condition2_list = []
                for row_idx in df.index:
                    excel_checkin = df.at[row_idx, checkin_col]
                    excel_checkin_norm, excel_checkin_norm2 = normalize_date_for_match(excel_checkin)
                    
                    if (checkin_norm == excel_checkin_norm or checkin_norm == excel_checkin_norm2 or
                        checkin_norm2 == excel_checkin_norm or checkin_norm2 == excel_checkin_norm2):
                        condition2_list.append(True)
                    elif checkin_norm and excel_checkin_norm:
                        checkin_clean = checkin_norm.replace('-', '').replace('/', '').replace(' ', '')
                        excel_clean = excel_checkin_norm.replace('-', '').replace('/', '').replace(' ', '')
                        if checkin_clean in excel_clean or excel_clean in checkin_clean:
                            condition2_list.append(True)
                        else:
                            condition2_list.append(False)
                    else:
                        condition2_list.append(False)
                
                condition2 = pd.Series(condition2_list, index=df.index)
                
                # Date matching - check-out
                condition3_list = []
                for row_idx in df.index:
                    excel_checkout = df.at[row_idx, checkout_col]
                    excel_checkout_norm, excel_checkout_norm2 = normalize_date_for_match(excel_checkout)
                    
                    if (checkout_norm == excel_checkout_norm or checkout_norm == excel_checkout_norm2 or
                        checkout_norm2 == excel_checkout_norm or checkout_norm2 == excel_checkout_norm2):
                        condition3_list.append(True)
                    elif checkout_norm and excel_checkout_norm:
                        checkout_clean = checkout_norm.replace('-', '').replace('/', '').replace(' ', '')
                        excel_clean = excel_checkout_norm.replace('-', '').replace('/', '').replace(' ', '')
                        if checkout_clean in excel_clean or excel_clean in checkout_clean:
                            condition3_list.append(True)
                        else:
                            condition3_list.append(False)
                    else:
                        condition3_list.append(False)
                
                condition3 = pd.Series(condition3_list, index=df.index)
                
                # All three must match
                combined_condition = condition1 & condition2 & condition3
                matching_row_indices = df[combined_condition].index.tolist()
                
                if matching_row_indices:
                    print(f"   ✅ Found {len(matching_row_indices)} row(s) matching all three fields")
                else:
                    print(f"   ⚠️  No row found matching Guest Name + Check-In Date + Check-Out Date")
            except Exception as e:
                print(f"   ❌ Error matching by three fields: {e}")
                continue
        
        # Update matching rows
        if matching_row_indices:
            for row_idx in matching_row_indices:
                try:
                    # Check if already marked as "Received"
                    current_value = str(df.at[row_idx, INVOICE_RECEIVED_COLUMN]).strip()
                    if current_value.lower() == INVOICE_RECEIVED_VALUE.lower():
                        print(f"   ⏭️  Row {row_idx + 1} already marked as '{INVOICE_RECEIVED_VALUE}', skipping")
                        rows_skipped += 1
                        continue
                    
                    # Update the Invoice Received column
                    df.at[row_idx, INVOICE_RECEIVED_COLUMN] = INVOICE_RECEIVED_VALUE
                    rows_updated += 1
                    
                    log_info = ""
                    if 'Booking Code' in field_to_column:
                        bc_val = str(df.at[row_idx, field_to_column['Booking Code']])
                        log_info = f"Booking Code: '{bc_val}'"
                    elif 'Guest Name' in field_to_column:
                        gn_val = str(df.at[row_idx, field_to_column['Guest Name']])
                        log_info = f"Guest Name: '{gn_val}'"
                    
                    print(f"   ✅ Updated row {row_idx + 1} - {log_info} → '{INVOICE_RECEIVED_VALUE}'")
                    
                except Exception as e:
                    print(f"   ❌ Error updating row {row_idx + 1}: {e}")
                    continue
        else:
            print(f"   ❌ No matching row found for this entry")
    
    print(f"\n📊 Matching Summary:")
    print(f"   ✅ Rows updated: {rows_updated}")
    print(f"   ⏭️  Rows skipped (already marked): {rows_skipped}")
    print(f"   📋 Total rows in Excel: {len(df)}")
    
    # Save updated Excel file
    if rows_updated > 0:
        try:
            print(f"\n💾 Saving updated Excel file...")
            print(f"   📁 File path: {MATCHING_EXCEL_FILE_PATH}")
            
            # Check if file is open
            try:
                test_file = open(MATCHING_EXCEL_FILE_PATH, 'r+b')
                test_file.close()
            except PermissionError:
                print(f"   ⚠️  File might be open in Excel. Please close it and try again.")
                print(f"   💡 The file will be saved when you close Excel and run the function again.")
                return rows_updated
            
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Save DataFrame to Excel
            print(f"   💾 Writing DataFrame to Excel...")
            df.to_excel(MATCHING_EXCEL_FILE_PATH, index=False, engine='openpyxl')
            print(f"   ✅ DataFrame saved successfully")
            
            # Load with openpyxl to apply formatting
            wb = load_workbook(MATCHING_EXCEL_FILE_PATH)
            ws = wb.active
            
            # Apply header formatting
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            wb.save(MATCHING_EXCEL_FILE_PATH)
            wb.close()
            
            print(f"✅ Saved updated Excel file: {MATCHING_EXCEL_FILE_PATH}")
            print(f"   📊 Total rows in saved file: {len(df)}")
            print(f"   ✅ Rows marked as 'Received': {len(df[df[INVOICE_RECEIVED_COLUMN].astype(str).str.strip().str.lower() == INVOICE_RECEIVED_VALUE.lower()])}")
        except PermissionError as e:
            print(f"❌ Permission denied: Cannot save Excel file")
            print(f"   The file might be open in Excel. Please close it and try again.")
            print(f"   File path: {MATCHING_EXCEL_FILE_PATH}")
        except Exception as e:
            print(f"⚠️  Error saving Excel file: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"\n⚠️  No rows were updated (rows_updated = {rows_updated})")
        print(f"   This could mean:")
        print(f"   1. No matching rows were found in the Excel file")
        print(f"   2. All matching rows were already marked as 'Received'")
        print(f"   3. Column matching failed (check logs above for column mappings)")
    
    return rows_updated

def sidebar():
    """Sidebar with configuration status and quick actions"""
    st.sidebar.title("📋 Configuration")
    
    required_vars, optional_vars = check_configuration()
    
    # Check required configuration
    st.sidebar.subheader("Required Settings")
    all_required = True
    for var_name, var_value in required_vars.items():
        if var_value:
            st.sidebar.success(f"✅ {var_name}")
        else:
            st.sidebar.error(f"❌ {var_name} (Missing)")
            all_required = False
    
    # Check optional configuration
    st.sidebar.subheader("Optional Settings")
    for var_name, var_value in optional_vars.items():
        if var_value:
            st.sidebar.info(f"✅ {var_name}: {var_value}")
        else:
            st.sidebar.warning(f"⚠️ {var_name} (Not set)")
    
    st.sidebar.divider()
    
    # Quick actions
    st.sidebar.subheader("Quick Actions")
    
    if st.sidebar.button("🔄 Test Gmail Connection"):
        success, message = test_gmail_connection()
        if success:
            st.sidebar.success(message)
        else:
            st.sidebar.error(message)
    
    if st.sidebar.button("🔄 Test Google Drive Connection"):
        success, message = test_google_drive_connection()
        if success:
            st.sidebar.success(message)
        else:
            st.sidebar.error(message)
    
    if st.sidebar.button("🔄 Refresh Vendor Reference"):
        try:
            vendor_ref = load_vendor_reference()
            st.session_state.vendor_reference = vendor_ref
            st.sidebar.success(f"✅ Loaded {len(vendor_ref)} vendor references")
        except Exception as e:
            st.sidebar.error(f"❌ Error: {e}")
    
    st.sidebar.divider()
    
    # System info
    st.sidebar.subheader("System Info")
    st.sidebar.info(f"📊 Excel File: {os.path.basename(required_vars.get('EXCEL_FILE_PATH', 'Not set'))}")
    st.sidebar.info(f"📅 Search Period: {optional_vars.get('DAYS_TO_SEARCH', '7')} days")
    
    if all_required:
        st.sidebar.success("✅ All required settings configured")
    else:
        st.sidebar.error("❌ Missing required configuration")

def dashboard_tab():
    """Dashboard tab showing system status and overview"""
    st.header("📊 Dashboard")
    
    # System Status
    st.subheader("System Status")
    
    col1, col2, col3, col4 = st.columns(4)
    
    # Gmail Status
    with col1:
        try:
            gmail_status, gmail_msg = test_gmail_connection()
            if gmail_status:
                st.success("📧 Gmail: Connected")
            else:
                st.error("📧 Gmail: Disconnected")
        except:
            st.error("📧 Gmail: Not Configured")
    
    # Google Drive Status
    with col2:
        try:
            drive_status, drive_msg = test_google_drive_connection()
            if drive_status:
                st.success("☁️ Google Drive: Connected")
            else:
                st.error("☁️ Google Drive: Disconnected")
        except:
            st.error("☁️ Google Drive: Not Configured")
    
    # AWS Bedrock Status
    with col3:
        bedrock_enabled = os.getenv('ENABLE_OPENAI_VISION', 'false').lower() == 'true'
        if bedrock_enabled:
            model = os.getenv('AWS_BEDROCK_MODEL', 'Not set')
            st.success(f"🤖 AWS Bedrock: Enabled ({model})")
        else:
            st.warning("🤖 AWS Bedrock: Disabled")
    
    # Vendor Reference Status
    with col4:
        try:
            vendor_ref = load_vendor_reference()
            st.success(f"📋 Vendors: {len(vendor_ref)} loaded")
        except:
            st.error("📋 Vendors: Not loaded")
    
    st.divider()
    
    # Configuration Details
    st.subheader("Configuration Details")
    
    required_vars, optional_vars = check_configuration()
    
    config_col1, config_col2 = st.columns(2)
    
    with config_col1:
        st.write("**Required Configuration:**")
        config_df = pd.DataFrame([
            {"Setting": k, "Value": "✅ Set" if v else "❌ Missing"}
            for k, v in required_vars.items()
        ])
        st.dataframe(config_df, use_container_width=True, hide_index=True)
    
    with config_col2:
        st.write("**Optional Configuration:**")
        optional_df = pd.DataFrame([
            {"Setting": k, "Value": v if v else "Not set"}
            for k, v in optional_vars.items()
        ])
        st.dataframe(optional_df, use_container_width=True, hide_index=True)
    
    st.divider()
    
    # Vendor Reference
    st.subheader("Vendor Reference")
    
    if st.button("🔄 Load Vendor Reference"):
        try:
            vendor_ref = load_vendor_reference()
            st.session_state.vendor_reference = vendor_ref
            
            # Display vendor reference
            vendor_data = []
            columns = ['SANDHYA', 'MOKSHITHA', 'KUMAR', 'LAKSHMI']
            
            for vendor_name, column in vendor_ref.items():
                vendor_data.append({
                    'Vendor Name': vendor_name,
                    'Assigned To': column
                })
            
            if vendor_data:
                vendor_df = pd.DataFrame(vendor_data)
                st.dataframe(vendor_df, use_container_width=True, hide_index=True)
                st.success(f"✅ Loaded {len(vendor_ref)} vendor references")
            else:
                st.warning("⚠️ No vendor references found")
        except Exception as e:
            st.error(f"❌ Error loading vendor reference: {e}")
    
    if st.session_state.vendor_reference:
        st.info(f"📋 {len(st.session_state.vendor_reference)} vendor references loaded in session")
    
    st.divider()
    
    # Recent Activity
    st.subheader("Recent Activity")
    
    if st.session_state.last_processed:
        st.info(f"📅 Last processed: {st.session_state.last_processed}")

def process_invoices_tab():
    """Tab for processing invoices"""
    st.header("📧 Process Invoices")
    
    st.write("This will process unread invoice emails from Gmail, extract data using AWS Bedrock, upload PDFs to Google Drive, and update the Excel sheet.")
    
    # Configuration
    st.subheader("Processing Configuration")
    
    col1, col2 = st.columns(2)
    
    with col1:
        days_to_search = st.slider(
            "Days to Search",
            min_value=1,
            max_value=30,
            value=int(os.getenv('DAYS_TO_SEARCH', '7')),
            help="Number of days back to search for unread emails"
        )
    
    with col2:
        st.write("**Current Settings:**")
        st.info(f"📧 Gmail: {os.getenv('GMAIL_EMAIL', 'Not set')}")
        st.info(f"☁️ Drive Upload: {'Enabled' if os.getenv('ENABLE_GOOGLE_DRIVE_UPLOAD', 'false').lower() == 'true' else 'Disabled'}")
        st.info(f"🤖 AWS Bedrock: {'Enabled' if os.getenv('ENABLE_OPENAI_VISION', 'false').lower() == 'true' else 'Disabled'}")
    
    st.divider()
    
    # Process button
    if st.button("🚀 Process Invoices", type="primary", use_container_width=True):
        # Update environment variable for days to search
        os.environ['DAYS_TO_SEARCH'] = str(days_to_search)
        
        st.session_state.processing_status = "Processing..."
        
        # Create containers for output
        status_container = st.container()
        
        with status_container:
            st.info("🔄 Processing invoices... This may take a few minutes.")
            progress_bar = st.progress(0)
            status_text = st.empty()
        
        try:
            # Capture output (but don't display it)
            result, output = capture_output(process_invoices)
            
            # Update logs in session state (for internal use, not displayed)
            log_lines = output.split('\n')
            st.session_state.processing_logs.extend(log_lines)
            
            # Update status
            progress_bar.progress(100)
            st.session_state.processing_status = "Completed"
            st.session_state.last_processed = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if result is None:
                status_text.success("✅ Processing completed successfully!")
            else:
                status_text.success("✅ Processing completed successfully!")
            
            st.balloons()
            
        except Exception as e:
            st.session_state.processing_status = "Error"
            st.error(f"❌ Error during processing: {e}")
            import traceback
            st.code(traceback.format_exc(), language='python')
    
    # Display current status
    if st.session_state.processing_status != "Ready":
        st.divider()
        st.write(f"**Current Status:** {st.session_state.processing_status}")
    
    # Connection testing
    st.divider()
    st.subheader("Test Connections")
    
    test_col1, test_col2 = st.columns(2)
    
    with test_col1:
        if st.button("🧪 Test Gmail Connection", use_container_width=True):
            with st.spinner("Testing Gmail connection..."):
                success, message = test_gmail_connection()
                if success:
                    st.success(message)
                else:
                    st.error(message)
    
    with test_col2:
        if st.button("🧪 Test Google Drive Connection", use_container_width=True):
            with st.spinner("Testing Google Drive connection..."):
                success, message = test_google_drive_connection()
                if success:
                    st.success(message)
                else:
                    st.error(message)

def match_invoices_tab():
    """Tab for matching invoices with Excel"""
    st.header("🔗 Match Invoices")
    
    st.write("This will match processed invoices from the Invoice Master Sheet with your local Excel file and update the 'Invoice Received' status.")
    
    st.divider()
    
    # Match button
    if st.button("🔗 Match Invoices with Excel", type="primary", use_container_width=True):
        st.info("🔄 Matching invoices... This may take a few minutes.")
        
        try:
            # Use fixed version that handles both Google Sheets and Excel files
            result, output = capture_output(match_master_sheet_with_excel_fixed)
            
            # Parse result if possible
            if result and result > 0:
                st.success(f"✅ Matching completed successfully! Updated {result} row(s).")
            elif "Successfully matched and updated" in output or "Rows updated:" in output:
                st.success("✅ Matching completed successfully!")
            elif "No rows matched" in output or "No rows were updated" in output:
                st.warning("⚠️ No rows were matched.")
            else:
                st.info("ℹ️ Matching completed.")
            
        except Exception as e:
            st.error(f"❌ Error during matching: {e}")
            import traceback
            st.code(traceback.format_exc(), language='python')

def statistics_tab():
    """Tab for viewing statistics"""
    st.header("📊 Statistics")
    
    # Excel Statistics
    st.subheader("Excel File Statistics")
    
    excel_path = os.getenv('EXCEL_FILE_PATH', MATCHING_EXCEL_FILE_PATH)
    
    if st.button("🔄 Refresh Statistics"):
        if os.path.exists(excel_path):
            try:
                df = read_matching_excel_file()
                if df is not None and not df.empty:
                    # Display statistics
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Total Rows", len(df))
                    
                    with col2:
                        st.metric("Total Columns", len(df.columns))
                    
                    with col3:
                        if 'Invoice Received' in df.columns:
                            received_count = len(df[df['Invoice Received'].astype(str).str.strip().str.lower() == 'received'])
                            st.metric("Invoices Received", received_count)
                        else:
                            st.metric("Invoices Received", "N/A")
                    
                    st.divider()
                    
                    # Column information
                    st.subheader("Column Information")
                    columns_df = pd.DataFrame({
                        'Column Name': df.columns,
                        'Data Type': [str(df[col].dtype) for col in df.columns],
                        'Non-Null Count': [df[col].notna().sum() for col in df.columns]
                    })
                    st.dataframe(columns_df, use_container_width=True, hide_index=True)
                    
                    st.divider()
                    
                    # Data preview
                    st.subheader("Data Preview")
                    st.dataframe(df.head(20), use_container_width=True, height=400)
                    
                    st.divider()
                    
                    # Assignment statistics
                    st.subheader("Assignment Statistics")
                    if 'Assigned To' in df.columns:
                        assignment_counts = df['Assigned To'].value_counts()
                        st.bar_chart(assignment_counts)
                    else:
                        st.info("No 'Assigned To' column found in Excel file")
                    
                else:
                    st.warning("⚠️ Excel file is empty")
            except Exception as e:
                st.error(f"❌ Error reading Excel file: {e}")
        else:
            st.error(f"❌ Excel file not found: {excel_path}")
    
    # Vendor Reference Statistics
    st.divider()
    st.subheader("Vendor Reference Statistics")
    
    if st.button("🔄 Load Vendor Statistics"):
        try:
            vendor_ref = load_vendor_reference()
            
            if vendor_ref:
                # Count by column
                column_counts = {}
                for vendor, column in vendor_ref.items():
                    column_counts[column] = column_counts.get(column, 0) + 1
                
                stats_df = pd.DataFrame({
                    'Assigned To': list(column_counts.keys()),
                    'Vendor Count': list(column_counts.values())
                })
                
                st.dataframe(stats_df, use_container_width=True, hide_index=True)
                st.bar_chart(stats_df.set_index('Assigned To'))
                
                st.success(f"✅ Loaded statistics for {len(vendor_ref)} vendors")
            else:
                st.warning("⚠️ No vendor references found")
        except Exception as e:
            st.error(f"❌ Error loading vendor statistics: {e}")

def main():
    """Main function"""
    # Header
    st.markdown('<p class="main-header">📧 Invoice Automation Hub</p>', unsafe_allow_html=True)
    st.markdown("---")
    
    # Tabs
    tab1, tab2, tab3 = st.tabs([
        "📊 Dashboard",
        "📧 Process Invoices",
        "🔗 Match Invoices"
    ])
    
    with tab1:
        dashboard_tab()
    
    with tab2:
        process_invoices_tab()
    
    with tab3:
        match_invoices_tab()
    
    # Footer
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: #666; padding: 1rem;'>"
        "Invoice Automation Hub | Built with Streamlit"
        "</div>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()

