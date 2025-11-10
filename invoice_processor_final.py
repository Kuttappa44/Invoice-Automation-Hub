#!/usr/bin/env python3
"""
Final Invoice Processor - Complete Flow
1. Check unread emails
2. Process invoices with AWS Bedrock
3. Upload PDFs to Google Drive folders
4. Update Excel sheet with extracted data
"""

import imaplib
import email
import re
import os
from datetime import datetime, timedelta
from typing import Optional
import pandas as pd
from fuzzywuzzy import fuzz, process
from dotenv import load_dotenv
import PyPDF2
from google_drive_uploader import GoogleDriveUploader
from openai_vision_extractor import OpenAIPropertyExtractor
import io

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

# Load environment variables
load_dotenv()

# Excel file path for matching (same as email_to_excel_mapper.py)
MATCHING_EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH', r'C:\Users\ADMIN\Desktop\Unmapped report for Mapping.xlsx')

# Google Sheets file ID for Invoice Processor Master Data
INVOICE_MASTER_SHEET_ID = os.getenv('INVOICE_MASTER_SHEET_ID', '10LdcnknrQqtrXM-UqnmXQ3ZOCvTD06nd')

# Fields to extract for matching (same as email_to_excel_mapper.py)
MATCHING_FIELDS = [
    'Booking Code',
    'Guest Name',
    'Hotel Name',
    'Check-In Date',
    'Check-Out Date'
]

# Column name for invoice received status
INVOICE_RECEIVED_COLUMN = 'Invoice Received'
INVOICE_RECEIVED_VALUE = 'Received'

# Invoice keywords
INVOICE_KEYWORDS = [
    'invoice', 'bill', 'receipt', 'payment', 'booking', 'reservation',
    'confirmation', 'voucher', 'ticket', 'statement', 'charge',
    'hotel', 'accommodation', 'travel', 'booking id', 'reservation id',
    'pending', 'required', 'urgent', 'bills'
]

def load_vendor_reference():
    """Load the Excel reference sheet with vendor names"""
    try:
        df = pd.read_excel('userlist.xlsx')
        vendor_reference = {}
        columns = ['SANDHYA', 'MOKSHITHA', 'KUMAR', 'LAKSHMI']
        
        for col in columns:
            if col in df.columns:
                vendors = df[col].dropna().astype(str).tolist()
                for vendor in vendors:
                    if vendor.strip():
                        vendor_reference[vendor.strip().lower()] = col
                        vendor_reference[vendor.strip()] = col
        
        print(f"✅ Loaded {len(vendor_reference)} vendor references")
        return vendor_reference
    except Exception as e:
        print(f"❌ Error loading vendor reference: {e}")
        return {}

def is_invoice_email(subject, body):
    """Check if email is an invoice based on subject and body"""
    text = f"{subject} {body}".lower()
    return any(keyword in text for keyword in INVOICE_KEYWORDS)

def decode_email_subject(subject):
    """Decode email subject if it's encoded"""
    try:
        decoded = email.header.decode_header(subject)
        decoded_subject = ""
        for part, encoding in decoded:
            if isinstance(part, bytes):
                if encoding:
                    decoded_subject += part.decode(encoding)
                else:
                    decoded_subject += part.decode('utf-8', errors='ignore')
            else:
                decoded_subject += part
        return decoded_subject
    except:
        return subject

def extract_text_from_pdf(pdf_data):
    """Extract text content from PDF data"""
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_data))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"   ⚠️  Error extracting PDF text: {e}")
        return ""

def extract_attachment_info(email_message):
    """Extract attachment information and content from email"""
    attachments = []
    attachment_texts = []
    
    if email_message.is_multipart():
        for part in email_message.walk():
            if part.get_content_disposition() == 'attachment':
                filename = part.get_filename()
                content_type = part.get_content_type()
                payload = part.get_payload(decode=True)
                size = len(payload) if payload else 0
                
                attachment_info = {
                    'filename': filename or 'Unknown',
                    'content_type': content_type,
                    'size_bytes': size,
                    'payload': payload  # Include payload for AWS Bedrock analysis
                }
                
                # Extract text from PDF attachments
                if content_type == 'application/pdf' and payload:
                    print(f"   📄 Processing PDF: {filename}")
                    pdf_text = extract_text_from_pdf(payload)
                    if pdf_text:
                        attachment_info['extracted_text'] = pdf_text[:1000] + "..." if len(pdf_text) > 1000 else pdf_text
                        attachment_texts.append(pdf_text)
                        print(f"   ✅ Extracted {len(pdf_text)} characters from PDF")
                
                attachments.append(attachment_info)
    
    # Return empty list if no attachments, not ['None']
    return attachments if attachments else [], attachment_texts

def extract_vendor_name_improved(subject, from_email, body, pdf_analysis=None, pdf_data=None, openai_extractor=None):
    """Extract vendor name with improved logic"""
    # First, try to get hotel name from PDF analysis if available
    hotel_name = None
    if pdf_analysis and isinstance(pdf_analysis, dict):
        for pdf_key, pdf_data_item in pdf_analysis.items():
            # Ensure pdf_data_item is a dictionary
            if isinstance(pdf_data_item, dict) and pdf_data_item.get('hotel_name') and pdf_data_item['hotel_name'] != 'Not Found':
                hotel_name = pdf_data_item['hotel_name']
                break
    
    # If no hotel name from PDF, try AWS Bedrock
    if not hotel_name and pdf_data and openai_extractor and openai_extractor.enabled:
        try:
            hotel_name = openai_extractor.extract_property_name_from_pdf(pdf_data)
            if hotel_name and hotel_name != 'Not Found':
                print(f"   🤖 AWS Bedrock extracted hotel: {hotel_name}")
        except Exception as e:
            print(f"   ⚠️  AWS Bedrock error: {e}")
    
    # If we have a hotel name, use it
    if hotel_name and hotel_name != 'Not Found':
        return hotel_name
    
    # Fallback to email-based extraction
    text_to_search = f"{subject} {from_email} {body}".lower()
    
    # Look for common hotel/property indicators
    hotel_indicators = [
        'hotel', 'resort', 'inn', 'chalet', 'executive', 'apartments',
        'suites', 'palace', 'tower', 'plaza', 'central', 'grand',
        'premier', 'luxury', 'boutique', 'international', 'airport',
        'city', 'garden', 'park', 'view', 'heights', 'manor', 'villa',
        'house', 'lodge', 'court', 'square', 'mall', 'center', 'centre',
        'complex', 'building', 'towers'
    ]
    
    for indicator in hotel_indicators:
        if indicator in text_to_search:
            # Extract the word before the indicator
            pattern = r'(\w+)\s+' + indicator
            match = re.search(pattern, text_to_search)
            if match:
                potential_name = match.group(1).title() + ' ' + indicator.title()
                return potential_name
    
    # If no hotel name found, return generic
    return 'Unknown Vendor'

def extract_email_body_full(email_message):
    """Extract plain text body and HTML body from email - returns both (similar to email_to_excel_mapper.py)"""
    body = ""
    html_body = ""
    
    if email_message.is_multipart():
        for part in email_message.walk():
            content_type = part.get_content_type()
            if content_type == "text/plain":
                try:
                    payload = part.get_payload(decode=True)
                    if payload:
                        plain_text = payload.decode('utf-8', errors='ignore')
                        body = body + "\n" + plain_text if body else plain_text
                except:
                    continue
            elif content_type == "text/html":
                try:
                    payload = part.get_payload(decode=True)
                    if payload:
                        html_part = payload.decode('utf-8', errors='ignore')
                        html_body = html_body + "\n" + html_part if html_body else html_part
                except:
                    continue
    else:
        try:
            payload = email_message.get_payload(decode=True)
            if payload:
                content = payload.decode('utf-8', errors='ignore')
                if '<html' in content.lower() or '<table' in content.lower():
                    html_body = content
                    body = re.sub(r'<[^>]+>', '', content)
                else:
                    body = content
        except:
            body = str(email_message.get_payload())
    
    if html_body and not body:
        body = re.sub(r'<[^>]+>', '', html_body)
    
    return body, html_body


def normalize_table_header(header_text):
    """Normalize HTML table header text to standardized field keys"""
    if not header_text:
        return None
    normalized = re.sub(r"\s+", " ", str(header_text)).strip().lower()
    normalized = normalized.replace("\xa0", " ")
    normalized = re.sub(r"[^a-z0-9\s_-]", "", normalized)
    normalized = normalized.replace("  ", " ")
    if not normalized:
        return None
    if ("booking" in normalized or "reservation" in normalized or "confirmation" in normalized or normalized.startswith("ref")) and any(
        token in normalized for token in ["code", "id", "number", "no", "ref", "confirmation", "booking"]
    ):
        return "booking_code"
    if "guest" in normalized and "name" in normalized:
        return "guest_name"
    if "client" in normalized and "name" in normalized:
        return "client_name"
    if "customer" in normalized and "name" in normalized:
        return "guest_name"
    if ("hotel" in normalized or "property" in normalized or "venue" in normalized) and "name" in normalized:
        return "property_name"
    if "check" in normalized and any(token in normalized for token in ["in", "arrival", "from", "start", "arr"]):
        return "check_in_date"
    if "check" in normalized and any(token in normalized for token in ["out", "departure", "to", "end", "dep"]):
        return "check_out_date"
    if "arrival" in normalized:
        return "check_in_date"
    if "departure" in normalized:
        return "check_out_date"
    if "guest" in normalized and any(token in normalized for token in ["count", "qty", "pax", "number"]):
        return "guest_count"
    if any(token in normalized for token in ["amount", "total", "grand total", "balance"]):
        return "total_amount"
    return None


def parse_html_tables_for_booking_data(html_body):
    """Parse HTML tables to extract booking information from structured layouts"""
    if not BeautifulSoup or not html_body:
        if not BeautifulSoup:
            print("   🛈 BeautifulSoup unavailable: skipping HTML table parsing")
        return []
    try:
        soup = BeautifulSoup(html_body, 'html.parser')
    except Exception as e:
        print(f"   ⚠️  HTML parsing failed: {e}")
        return []
    extracted_entries = []
    tables = soup.find_all('table')
    print(f"   🧾 Detected {len(tables)} HTML table(s) in email body")
    for table in tables:
        rows = []
        for row in table.find_all('tr'):
            cells = row.find_all(['th', 'td'])
            cell_texts = []
            for cell in cells:
                text = cell.get_text(separator=' ', strip=True)
                text = text.replace('\xa0', ' ').strip()
                if text:
                    cell_texts.append(text)
            if cell_texts:
                rows.append(cell_texts)
        if not rows:
            print("      ⚠️  Table skipped: no rows with text content")
            continue
        header_row = rows[0]
        header_map = [normalize_table_header(header) for header in header_row]
        print(f"      ↳ Raw headers: {header_row}")
        print(f"        Normalized headers: {header_map}")
        useful_headers = [h for h in header_map if h]
        if useful_headers and len(rows) > 1:
            for data_row in rows[1:]:
                entry = {}
                for idx, value in enumerate(data_row):
                    if idx >= len(header_map):
                        continue
                    key = header_map[idx]
                    if not key:
                        continue
                    if key not in entry or not entry[key]:
                        entry[key] = value.strip()
                if any(entry.get(k) for k in ['booking_code', 'guest_name', 'client_name', 'check_in_date', 'check_out_date']):
                    print(f"        ✓ Table row parsed: {entry}")
                    extracted_entries.append(entry)
                else:
                    print(f"        ⚠️  Table row ignored (no relevant fields): {data_row}")
        elif useful_headers:
            print("      ⚠️  Table has headers but no data rows")
        else:
            kv_entry = {}
            for row in rows:
                if len(row) != 2:
                    continue
                key = normalize_table_header(row[0])
                if key:
                    value = row[1].strip()
                    if value:
                        kv_entry[key] = value
            if kv_entry:
                print(f"        ✓ Key/value rows parsed: {kv_entry}")
                extracted_entries.append(kv_entry)
            else:
                print("      ⚠️  Key/value table yielded no usable data")
    return extracted_entries

def extract_booking_code_from_email(email_message, subject, body, html_body, openai_extractor, attachment_texts=None):
    """Extract booking details (code, guest, dates, property) from email content"""

    booking_details = {}

    def store_detail(key, value):
        if value and isinstance(value, str):
            value_clean = value.strip()
            if value_clean and key not in booking_details:
                booking_details[key] = value_clean

    def clean_booking_value(value: str):
        value = value.strip()
        value = re.sub(r'^[\-:>#~\s]+', '', value)
        value = re.sub(r'[\s,.;]+$', '', value)
        value = value.strip('"\'`')
        value = value.replace(' ', '')
        value = value.strip()
        digits_match = re.search(r'(\d{7})', value)
        if digits_match:
            return digits_match.group(1)
        return value

    def is_valid_booking_code(value: str, context_label: Optional[str] = None) -> bool:
        if not value:
            return False
        digits = re.sub(r'\D', '', value)
        if not digits:
            return False
        if len(digits) != 7:
            return False
        return digits.isdigit()

    booking_code = None

    # 1. Structured HTML table parsing (handles emails with booking tables)
    if html_body:
        print("   🔎 Parsing HTML tables for booking data...")
        table_entries = parse_html_tables_for_booking_data(html_body)
        if table_entries:
            print(f"   ✅ Parsed {len(table_entries)} table row(s) with potential data")
            booking_details['table_entries'] = table_entries
        for entry in table_entries:
            candidate = entry.get('booking_code') or entry.get('confirmation_number') or entry.get('reference')
            if candidate and not booking_code:
                candidate_clean = clean_booking_value(candidate)
                if is_valid_booking_code(candidate_clean, context_label='email_table'):
                    booking_code = candidate_clean
                    booking_details['booking_code_source'] = 'email_table'
                    print(f"   ✅ Booking code detected in table row: {booking_code}")
            if entry.get('guest_name'):
                store_detail('guest_name', entry['guest_name'])
            if entry.get('client_name'):
                store_detail('client_name', entry['client_name'])
                if 'guest_name' not in booking_details:
                    store_detail('guest_name', entry['client_name'])
            if entry.get('property_name'):
                store_detail('property_name', entry['property_name'])
            if entry.get('check_in_date'):
                store_detail('check_in_date', entry['check_in_date'])
            if entry.get('check_out_date'):
                store_detail('check_out_date', entry['check_out_date'])
        if booking_code:
            print(f"   ✅ Booking code extracted from HTML table: {booking_code}")
            booking_details['booking_code'] = booking_code
            return booking_code, booking_details
        elif table_entries:
            print("   ⚠️  Table rows found but no valid booking code parsed")
        else:
            print("   ⚠️  No booking-relevant tables detected in HTML body")

    # 2. AI extraction via AWS Bedrock
    ai_response = None
    if openai_extractor and openai_extractor.enabled:
        try:
            print(f"   🔍 Extracting booking code using AWS Bedrock...")
            ai_response = openai_extractor.extract_booking_details_from_email(body, html_body, subject)
            if ai_response:
                booking_details['ai_raw_response'] = ai_response
                booking_patterns = [
                    r'BOOKING\s+CODE[:\s]+([A-Z0-9\-]+)',
                    r'BOOKING\s+ID[:\s]+([A-Z0-9\-]+)',
                    r'BOOKING\s+REFERENCE[:\s]+([A-Z0-9\-]+)',
                    r'CONFIRMATION\s+NUMBER[:\s]+([A-Z0-9\-]+)',
                    r'BOOKING\s+NUMBER[:\s]+([A-Z0-9\-]+)',
                    r'BOOKING[:\s]+([A-Z0-9\-]+)',
                    r'(?:REF|REFERENCE|REF\s+NO)[:\s]+([A-Z0-9\-]+)',
                    r'\bBOOKING[^\d]{0,40}(\d{7})'
                ]
                for pattern in booking_patterns:
                    match = re.search(pattern, ai_response, re.IGNORECASE)
                    if match:
                        candidate = clean_booking_value(match.group(1))
                        if is_valid_booking_code(candidate, context_label='bedrock'):
                            booking_code = candidate
                            booking_details['booking_code_source'] = 'bedrock'
                            break
                if not booking_code:
                    lines = ai_response.split('\n')
                    for line in lines:
                        line = line.strip()
                        if not line:
                            continue
                        if re.match(r'^BOOKING\s+(?:CODE|ID|REFERENCE|NUMBER)[:\s]+', line, re.IGNORECASE):
                            colon_index = line.find(':')
                            if colon_index >= 0:
                                candidate = clean_booking_value(line[colon_index + 1:])
                                if is_valid_booking_code(candidate, context_label='bedrock_line'):
                                    booking_code = candidate
                                    booking_details['booking_code_source'] = 'bedrock'
                                    break
                if booking_code:
                    booking_details['booking_code'] = booking_code
                    print(f"   ✅ Booking code extracted via AWS Bedrock: {booking_code}")
                else:
                    print("   ⚠️  AWS Bedrock did not return a usable booking code")
        except Exception as e:
            print(f"   ⚠️  Error extracting booking code from email via AI: {e}")

    # 3. Fallback manual parsing if AI/table extraction missed it
    if not booking_code:
        print("   🔁 Entering fallback booking code scan...")
        try:
            fallback_sections = [subject or '', body or '']
            table_strings = []

            if html_body:
                text_from_html = re.sub(r'<[^>]+>', ' ', html_body)
                fallback_sections.append(text_from_html)
                if BeautifulSoup:
                    try:
                        soup = BeautifulSoup(html_body, 'html.parser')
                        linearized_html_text = soup.get_text(separator='\n', strip=True)
                        if linearized_html_text:
                            fallback_sections.append(linearized_html_text)
                        for table in soup.find_all('table'):
                            for row in table.find_all('tr'):
                                cells = row.find_all(['th', 'td'])
                                row_text = ' '.join(cell.get_text(separator=' ', strip=True) for cell in cells)
                                if row_text:
                                    table_strings.append(row_text)
                    except Exception as soup_error:
                        print(f"   ⚠️  Error parsing HTML tables for fallback booking code: {soup_error}")
                else:
                    print("   ⚠️  BeautifulSoup not available; skipping table-specific fallback extraction")

            try:
                alternative_parts = []
                if email_message.is_multipart():
                    for part in email_message.walk():
                        content_type = part.get_content_type() or ''
                        if content_type == 'text/plain':
                            payload = part.get_payload(decode=True)
                            if payload:
                                alternative_parts.append(payload.decode('utf-8', errors='ignore'))
                        elif content_type == 'text/html':
                            payload = part.get_payload(decode=True)
                            if payload:
                                alternative_parts.append(re.sub(r'<[^>]+>', ' ', payload.decode('utf-8', errors='ignore')))
                else:
                    payload = email_message.get_payload(decode=True)
                    if payload:
                        decoded = payload.decode('utf-8', errors='ignore')
                        if '<html' in decoded.lower():
                            alternative_parts.append(re.sub(r'<[^>]+>', ' ', decoded))
                        else:
                            alternative_parts.append(decoded)
                if alternative_parts:
                    fallback_sections.extend(alternative_parts)
            except Exception as inner_e:
                print(f"   ⚠️  Error collecting additional email parts for booking code: {inner_e}")

            if attachment_texts:
                fallback_sections.extend(attachment_texts)

            if table_strings:
                fallback_sections.extend(table_strings)

            combined_text = '\n'.join(section for section in fallback_sections if section)
            combined_text = combined_text.replace('\r', ' ')

            if combined_text.strip():
                print(f"   📄 Combined fallback text length: {len(combined_text)} characters")
                fallback_patterns = [
                    r'\bBOOKING\s*(?:CODE|ID|REFERENCE|REF|NO|NUMBER)?\s*(?:[:#\-]|NO\.\s*)?\s*([A-Za-z0-9]{3,}[A-Za-z0-9\-/]{0,})',
                    r'\bCONFIRMATION\s*(?:NUMBER|NO|#)?\s*(?:[:#\-]|NO\.\s*)?\s*([A-Za-z0-9]{3,}[A-Za-z0-9\-/]{0,})',
                    r'\bRESERVATION\s*(?:NUMBER|NO|ID|CODE)?\s*(?:[:#\-]|NO\.\s*)?\s*([A-Za-z0-9]{3,}[A-Za-z0-9\-/]{0,})',
                    r'\bREFERENCE\s*(?:NO|NUMBER|ID)?\s*(?:[:#\-]|NO\.\s*)?\s*([A-Za-z0-9]{3,}[A-Za-z0-9\-/]{0,})',
                    r'\bCONF\s*(?:NO|NUMBER|ID)?\s*(?:[:#\-]|NO\.\s*)?\s*([A-Za-z0-9]{3,}[A-Za-z0-9\-/]{0,})',
                    r'\bBOOKING[^\d]{0,40}(\d{7})'
                ]

                for pattern in fallback_patterns:
                    match = re.search(pattern, combined_text, re.IGNORECASE)
                    if match:
                        candidate = clean_booking_value(match.group(1))
                        if is_valid_booking_code(candidate):
                            booking_code = candidate
                            booking_details['booking_code_source'] = 'regex_fallback'
                            print(f"   ✅ Regex fallback matched booking code: {booking_code}")
                            break

                if not booking_code:
                    keyword_window = 120
                    keywords = ['booking', 'reference', 'confirmation', 'reservation', 'ref', 'conf']
                    for keyword in keywords:
                        keyword_matches = list(re.finditer(keyword, combined_text, re.IGNORECASE))
                        if keyword_matches:
                            print(f"   🔍 Keyword '{keyword}' occurrences: {len(keyword_matches)}")
                        for match in keyword_matches:
                            start = max(0, match.start() - keyword_window)
                            end = min(len(combined_text), match.end() + keyword_window)
                            window_text = combined_text[start:end]
                            code_match = re.search(r'(\d{7})', window_text)
                            if code_match:
                                candidate = clean_booking_value(code_match.group(1))
                                if is_valid_booking_code(candidate, context_label=keyword):
                                    booking_code = candidate
                                    booking_details['booking_code_source'] = 'heuristic_window'
                                    print(f"   ✅ Heuristic window ({keyword}) found booking code: {booking_code}")
                                    break
                        if booking_code:
                            break

            if booking_code:
                booking_details['booking_code'] = booking_code
                print(f"   ✅ Booking code extracted via fallback analysis: {booking_code}")
            else:
                print("   ⚠️  Fallback scan exhausted without finding a booking code")
        except Exception as e:
            print(f"   ⚠️  Fallback booking code extraction error: {e}")

    if booking_code and 'booking_code' not in booking_details:
        booking_details['booking_code'] = booking_code

    return booking_code, booking_details

def extract_comprehensive_email_data(email_message, subject, from_email, body, openai_extractor=None):
    """Extract comprehensive information from email and attachments"""
    # Extract email headers
    to_email = email_message.get('To', '')
    cc_email = email_message.get('Cc', '')
    bcc_email = email_message.get('Bcc', '')
    reply_to = email_message.get('Reply-To', '')
    message_id = email_message.get('Message-ID', '')
    in_reply_to = email_message.get('In-Reply-To', '')
    references = email_message.get('References', '')
    
    # Extract both plain text and HTML body for booking code extraction
    body_full, html_body = extract_email_body_full(email_message)
    if not body_full:
        body_full = body  # Fallback to provided body
    
    # Extract attachments
    attachments, attachment_texts = extract_attachment_info(email_message)
    
    # Extract booking code and related details from email/tables/AI
    booking_code = None
    email_booking_details = {}
    try:
        booking_code, email_booking_details = extract_booking_code_from_email(
            email_message,
            subject,
            body_full,
            html_body,
            openai_extractor,
            attachment_texts=attachment_texts
        )
    except Exception as e:
        print(f"   ⚠️  Error during booking detail extraction: {e}")
        booking_code = None
        email_booking_details = {}
    
    # Process PDF attachments with AWS Bedrock
    pdf_analysis = {}
    if attachments and len(attachments) > 0:
        # Filter out any non-dict items (like 'None' strings)
        valid_attachments = [a for a in attachments if isinstance(a, dict)]
        
        if not valid_attachments:
            print(f"   ⚠️  No valid attachments found (all are non-dictionary types)")
        else:
            for i, attachment in enumerate(valid_attachments):
                if attachment.get('content_type') == 'application/pdf' and attachment.get('payload'):
                    print(f"   🔍 Analyzing PDF {i+1} with AWS Bedrock...")
                    
                    # Extract text for fallback
                    pdf_text = extract_text_from_pdf(attachment['payload'])
                    
                    # Try AWS Bedrock first
                    openai_data = None
                    if openai_extractor and openai_extractor.enabled:
                        try:
                            openai_data = openai_extractor.extract_comprehensive_invoice_data_from_pdf(attachment['payload'])
                            if openai_data:
                                print(f"   ✅ AWS Bedrock extracted data for PDF {i+1}")
                        except Exception as e:
                            print(f"   ⚠️  AWS Bedrock error for PDF {i+1}: {e}")
                    
                    # Create PDF data structure
                    pdf_data = {
                        'hotel_name': 'Not Found',
                        'guest_name': 'Not Found',
                        'bill_number': 'Not Found',
                        'bill_date': 'Not Found',
                        'arrival_date': 'Not Found',
                        'departure_date': 'Not Found',
                        'room_number': 'Not Found',
                        'number_of_pax': 'Not Found',
                        'total_amount': 'Not Found',
                        'gst_number': 'Not Found',
                        'pan_number': 'Not Found',
                        'hotel_address': 'Not Found',
                        'hotel_phone': 'Not Found',
                        'hotel_email': 'Not Found',
                        'filename': attachment.get('filename', f'pdf_{i+1}'),
                        'payload': attachment.get('payload')  # Keep PDF data for upload
                    }
                    
                    # Use AWS Bedrock extracted data if available
                    if openai_data:
                        pdf_data['bedrock_extracted_data'] = openai_data
                        # Parse AWS Bedrock response to extract structured data
                        if 'HOTEL:' in openai_data:
                            hotel_match = re.search(r'HOTEL:\s*([^\n]+)', openai_data)
                            if hotel_match:
                                pdf_data['hotel_name'] = hotel_match.group(1).strip()
                        
                        if 'GUEST:' in openai_data:
                            guest_match = re.search(r'GUEST:\s*([^\n]+)', openai_data)
                            if guest_match:
                                pdf_data['guest_name'] = guest_match.group(1).strip()
                        
                        # Note: Booking code is extracted from email, not PDF (see extract_booking_code_from_email)
                        
                        if 'BILL NO:' in openai_data:
                            bill_match = re.search(r'BILL NO:\s*([^\n]+)', openai_data)
                            if bill_match:
                                pdf_data['bill_number'] = bill_match.group(1).strip()
                        
                        if 'AMOUNT:' in openai_data:
                            amount_match = re.search(r'AMOUNT:\s*([^\n]+)', openai_data)
                            if amount_match:
                                pdf_data['total_amount'] = amount_match.group(1).strip()
                        
                        if 'ROOM:' in openai_data:
                            room_match = re.search(r'ROOM:\s*([^\n]+)', openai_data)
                            if room_match:
                                pdf_data['room_number'] = room_match.group(1).strip()
                        
                        if 'GUESTS:' in openai_data:
                            guests_match = re.search(r'GUESTS:\s*([^\n]+)', openai_data)
                            if guests_match:
                                pdf_data['number_of_pax'] = guests_match.group(1).strip()
                        
                        if 'CHECK-IN:' in openai_data:
                            checkin_match = re.search(r'CHECK-IN:\s*([^\n]+)', openai_data)
                            if checkin_match:
                                pdf_data['arrival_date'] = checkin_match.group(1).strip()
                        
                        if 'CHECK-OUT:' in openai_data:
                            checkout_match = re.search(r'CHECK-OUT:\s*([^\n]+)', openai_data)
                            if checkout_match:
                                pdf_data['departure_date'] = checkout_match.group(1).strip()
                        
                        if 'BILL DATE:' in openai_data:
                            billdate_match = re.search(r'BILL DATE:\s*([^\n]+)', openai_data)
                            if billdate_match:
                                pdf_data['bill_date'] = billdate_match.group(1).strip()
                        
                        if 'GST:' in openai_data:
                            gst_match = re.search(r'GST:\s*([^\n]+)', openai_data)
                            if gst_match:
                                pdf_data['gst_number'] = gst_match.group(1).strip()
                    else:
                        print(f"   ❌ PDF couldn't be processed")
                        pdf_data['hotel_name'] = 'PDF couldn\'t be processed'
                    
                    pdf_analysis[f'pdf_{i+1}'] = pdf_data
    
    # Basic text extraction for email metadata only
    combined_text = body_full + ' ' + subject
    amounts = extract_amounts(combined_text)
    urgency_level = extract_urgency_level(combined_text)
    
    result = {
        'to_email': to_email,
        'cc_email': cc_email,
        'bcc_email': bcc_email,
        'reply_to': reply_to,
        'message_id': message_id,
        'in_reply_to': in_reply_to,
        'references': references,
        'attachments': attachments,
        'attachment_texts': attachment_texts,
        'pdf_analysis': pdf_analysis,
        'amounts': amounts,
        'urgency_level': urgency_level,
        'booking_code': booking_code,  # Booking code extracted from email, not PDF
        'email_booking_details': email_booking_details
    }

    if email_booking_details:
        if email_booking_details.get('guest_name'):
            result['email_guest_name'] = email_booking_details['guest_name']
        if email_booking_details.get('client_name'):
            result['email_client_name'] = email_booking_details['client_name']
        if email_booking_details.get('property_name'):
            result['email_property_name'] = email_booking_details['property_name']
        if email_booking_details.get('check_in_date'):
            result['email_check_in_date'] = email_booking_details['check_in_date']
        if email_booking_details.get('check_out_date'):
            result['email_check_out_date'] = email_booking_details['check_out_date']

    return result

def extract_amounts(text):
    """Extract comprehensive monetary amounts from text"""
    amount_patterns = [
        # Currency symbols with amounts
        r'[\$₹€£]\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*[\$₹€£]',
        
        # Indian Rupee patterns
        r'₹\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'Rs\.?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'INR\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        
        # Generic number patterns that could be amounts
        r'(?<!\d)(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)(?!\d)',
    ]
    
    amounts = []
    for pattern in amount_patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                # Clean and convert to float
                clean_amount = match.replace(',', '')
                amount = float(clean_amount)
                # Filter out very small amounts that are likely not real prices
                if amount >= 1.0:
                    amounts.append(amount)
            except ValueError:
                continue
    
    return list(set(amounts))  # Remove duplicates

def extract_urgency_level(text):
    """Extract urgency level from text"""
    urgency_keywords = {
        'URGENT': ['urgent', 'asap', 'immediately', 'rush', 'priority'],
        'HIGH': ['high priority', 'important', 'critical'],
        'NORMAL': ['normal', 'standard', 'regular']
    }
    
    text_lower = text.lower()
    for level, keywords in urgency_keywords.items():
        if any(keyword in text_lower for keyword in keywords):
            return level
    
    return 'NORMAL'

def find_vendor_column(vendor_name, vendor_reference):
    """Find the appropriate column for a vendor using fuzzy matching"""
    if not vendor_name or vendor_name == 'Unknown Vendor':
        return 'UNASSIGNED'
    
    print(f"   🔍 Looking for vendor: '{vendor_name}'")
    
    # Direct match first (exact match)
    if vendor_name.lower() in vendor_reference:
        result = vendor_reference[vendor_name.lower()]
        print(f"   ✅ Direct match found: {result}")
        return result
    
    # Try partial matching for hotel names
    vendor_lower = vendor_name.lower()
    best_partial_match = None
    best_partial_score = 0
    
    for ref_name, column in vendor_reference.items():
        ref_lower = ref_name.lower()
        
        # Check if any significant word from the hotel name matches
        hotel_words = vendor_lower.split()
        ref_words = ref_lower.split()
        
        # Check for significant word matches (words longer than 3 characters)
        significant_matches = 0
        for hotel_word in hotel_words:
            if len(hotel_word) > 3:  # Only consider significant words
                for ref_word in ref_words:
                    if len(ref_word) > 3:
                        # Check if words are similar (fuzzy match)
                        if fuzz.ratio(hotel_word, ref_word) > 80:
                            significant_matches += 1
                            break
        
        # Also check for partial string matches (for cases like "Sapphire" in "SSAPPHIRE")
        partial_match_score = 0
        for hotel_word in hotel_words:
            if len(hotel_word) > 4:  # Only consider longer words
                for ref_word in ref_words:
                    if len(ref_word) > 4:
                        # Check if one word contains the other or vice versa
                        if hotel_word in ref_word or ref_word in hotel_word:
                            partial_match_score += 1
                        # Also check fuzzy match for partial words
                        elif fuzz.ratio(hotel_word, ref_word) > 70:
                            partial_match_score += 0.5
        
        # Calculate total score
        total_score = significant_matches + partial_match_score
        
        if total_score > best_partial_score:
            best_partial_score = total_score
            best_partial_match = (ref_name, column)
        
        # If we have at least 2 significant word matches, consider it a match
        if significant_matches >= 2:
            print(f"   ✅ Partial match found: '{ref_name}' -> {column} ({significant_matches} words match)")
            return column
    
    # If we have a good partial match, use it
    if best_partial_match and best_partial_score >= 1.5:
        ref_name, column = best_partial_match
        print(f"   ✅ Best partial match found: '{ref_name}' -> {column} (score: {best_partial_score})")
        return column
    
    # Fuzzy matching on the full name
    best_match = process.extractOne(vendor_name, list(vendor_reference.keys()), scorer=fuzz.ratio)
    if best_match and best_match[1] >= 60:  # Lowered threshold to 60%
        print(f"   ✅ Fuzzy match found: '{best_match[0]}' -> {vendor_reference[best_match[0]]} ({best_match[1]}% match)")
        return vendor_reference[best_match[0]]
    
    print(f"   ❌ No match found for: '{vendor_name}'")
    return 'UNASSIGNED'

def get_drive_folder_name(excel_column):
    """Convert Excel column name to Google Drive folder name"""
    # Now MOKSHITHA maps directly to MOKSHITHA folder
    return excel_column

def upload_pdf_to_drive(pdf_data, assigned_column, drive_uploader):
    """Upload PDF file to Google Drive folder"""
    if not drive_uploader or not pdf_data.get('payload'):
        return None
    
    try:
        # Create safe filename with date prefix
        original_filename = pdf_data.get('filename', 'unknown.pdf')
        safe_filename = re.sub(r'[^\w\s\-\.]', '', original_filename).strip()
        safe_filename = safe_filename.replace(' ', '_')
        
        # Add date prefix in YYYYMMDD format
        processing_date = datetime.now().strftime('%Y%m%d')
        date_filename = f"{processing_date}_{safe_filename}"
        
        # Upload to Google Drive - convert Excel column to Drive folder name
        folder_name = get_drive_folder_name(assigned_column) if assigned_column != 'Not Found' else 'UNASSIGNED'
        print(f"   🔧 DEBUG: assigned_column='{assigned_column}', folder_name='{folder_name}'")
        print(f"   📅 Original: {original_filename} -> Date format: {date_filename}")
        file_id = drive_uploader.upload_pdf_data(pdf_data['payload'], date_filename, folder_name)
        
        if file_id:
            print(f"   ☁️  Uploaded PDF to Google Drive: {folder_name}/{date_filename}")
            return file_id
        else:
            print(f"   ⚠️  Failed to upload PDF to Google Drive")
            return None
            
    except Exception as e:
        print(f"   ❌ Error uploading PDF to Drive: {e}")
        return None

def create_or_get_excel_file(drive_uploader):
    """Create a new Excel file in Google Drive or get existing one"""
    master_file_id = os.getenv('MASTER_EXCEL_FILE_ID', None)
    
    # Try to access existing file if ID is provided
    if master_file_id:
        try:
            file_metadata = drive_uploader.service.files().get(fileId=master_file_id, fields='id,name').execute()
            print(f"📊 Found existing Excel file: {file_metadata.get('name', 'Unknown')}")
            return master_file_id
        except Exception as e:
            print(f"⚠️  Cannot access existing file, creating new one...")
    
    # Create a new Excel file
    print("📊 Creating new Excel file in Google Drive...")
    try:
        # Create a new Excel file with headers
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Data"
        
        # Define headers
        headers = [
            'S.No', 'Processing Date', 'Email From', 'Email Subject', 'Assigned To',
            'Hotel Name', 'Guest Name', 'Booking Code', 'Bill Number', 'Bill Date', 'Room Number',
            'Number of Guests', 'Check-in Date', 'Check-out Date', 'Total Amount',
            'GST Number', 'PDF Filename', 'Drive File ID', 'Urgency Level'
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        column_widths = [8, 20, 25, 40, 15, 35, 25, 18, 18, 15, 12, 15, 15, 15, 18, 20, 25, 25, 12]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        
        # Save to temporary file
        temp_file = 'temp_new_master.xlsx'
        wb.save(temp_file)
        wb.close()  # Close the workbook to release the file
        
        # Upload to Google Drive
        from googleapiclient.http import MediaFileUpload
        file_metadata = {
            'name': 'Invoice Processor Master Data.xlsx',
            'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        media = MediaFileUpload(temp_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        file = drive_uploader.service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id,name,webViewLink,webContentLink'
        ).execute()
        
        new_file_id = file.get('id')
        
        # Make file shareable (anyone with link can view)
        permission = {
            'type': 'anyone',
            'role': 'reader'
        }
        drive_uploader.service.permissions().create(
            fileId=new_file_id,
            body=permission
        ).execute()
        
        # Get shareable link
        shareable_link = f"https://docs.google.com/spreadsheets/d/{new_file_id}/edit"
        
        print(f"✅ Created new Excel file: {file.get('name')}")
        print(f"🔗 Shareable link: {shareable_link}")
        print(f"📋 File ID: {new_file_id}")
        print(f"💡 Add this to your .env file: MASTER_EXCEL_FILE_ID={new_file_id}")
        
        # Clean up temp file - ensure it's closed first
        try:
            import time
            time.sleep(0.5)  # Brief wait to ensure file is fully closed
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except PermissionError as e:
            print(f"⚠️  Could not delete temp file (may be in use): {e}")
            # Not critical, continue anyway
        
        return new_file_id
        
    except Exception as e:
        print(f"❌ Error creating new Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None

def update_excel_sheet(invoice_data_list, drive_uploader):
    """Update the master Excel sheet with new invoice data"""
    if not invoice_data_list:
        print("⚠️  No invoice data to save")
        return None
    
    # Get or create Excel file
    master_file_id = create_or_get_excel_file(drive_uploader)
    if not master_file_id:
        print("❌ Could not create or access Excel file")
        return None
    
    try:
        # Download current file
        file_content = drive_uploader.service.files().get_media(fileId=master_file_id).execute()
        with open('temp_master.xlsx', 'wb') as f:
            f.write(file_content)
        print("📊 Downloaded current Excel file")
        
        # Load existing data
        df = pd.read_excel('temp_master.xlsx')
        print(f"📊 Current Excel has {len(df)} rows")
        
        # Prepare new data to append
        new_rows = []
        for invoice in invoice_data_list:
            pdf_analysis = invoice.get('pdf_analysis', {})
            # Ensure pdf_analysis is a dictionary
            if not isinstance(pdf_analysis, dict):
                print(f"   ⚠️  Skipping invoice - pdf_analysis is not a dictionary (type: {type(pdf_analysis)})")
                continue
            
            for pdf_key, pdf_data in pdf_analysis.items():
                # Ensure pdf_data is a dictionary
                if not isinstance(pdf_data, dict):
                    print(f"   ⚠️  Skipping PDF {pdf_key} - not a dictionary (type: {type(pdf_data)})")
                    continue
                
                # Skip if PDF couldn't be processed
                if pdf_data.get('hotel_name') == 'PDF couldn\'t be processed':
                    continue
                
                # Create new row data with proper date formatting
                processing_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                email_from = invoice.get('from_email', '').replace('<', '').replace('>', '')
                email_subject = invoice.get('subject', '')[:50] + '...' if len(invoice.get('subject', '')) > 50 else invoice.get('subject', '')
                assigned_folder = invoice.get('assigned_column', 'UNASSIGNED')
                hotel_name = pdf_data.get('hotel_name', 'Not Found')
                guest_name = pdf_data.get('guest_name', 'Not Found')
                email_details = invoice.get('email_booking_details', {}) or {}
                email_guest_name = invoice.get('email_guest_name') or email_details.get('guest_name') or email_details.get('client_name') or invoice.get('email_client_name')
                email_property_name = invoice.get('email_property_name') or email_details.get('property_name')
                email_check_in = invoice.get('email_check_in_date') or email_details.get('check_in_date')
                email_check_out = invoice.get('email_check_out_date') or email_details.get('check_out_date')
                if (not hotel_name or hotel_name in ['Not Found', '']) and email_property_name:
                    hotel_name = email_property_name
                if (not guest_name or guest_name in ['Not Found', '']) and email_guest_name:
                    guest_name = email_guest_name
                # Get booking code from email (not PDF) - extracted using email_to_excel_mapper.py logic
                booking_code = invoice.get('booking_code', 'Not Found')
                if not booking_code or booking_code == 'None':
                    booking_code = 'Not Found'
                bill_number = pdf_data.get('bill_number', 'Not Found')
                
                # Format bill date properly
                bill_date = pdf_data.get('bill_date', 'Not Found')
                if bill_date != 'Not Found' and bill_date:
                    try:
                        # Try to parse and reformat the date
                        if '/' in str(bill_date):
                            # Handle DD/MM/YY or DD/MM/YYYY format
                            date_parts = str(bill_date).split('/')
                            if len(date_parts) == 3:
                                day, month, year = date_parts
                                if len(year) == 2:
                                    year = '20' + year
                                bill_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    except:
                        pass  # Keep original if parsing fails
                
                room_number = pdf_data.get('room_number', 'Not Found')
                number_of_guests = pdf_data.get('number_of_pax', 'Not Found')
                if (not number_of_guests or number_of_guests in ['Not Found', '']) and email_details.get('guest_count'):
                    number_of_guests = email_details.get('guest_count')
                
                # Format check-in and check-out dates properly
                check_in_date = pdf_data.get('arrival_date', 'Not Found')
                if (not check_in_date or check_in_date == 'Not Found') and email_check_in:
                    check_in_date = email_check_in
                if check_in_date != 'Not Found' and check_in_date:
                    try:
                        if '/' in str(check_in_date):
                            date_parts = str(check_in_date).split('/')
                            if len(date_parts) == 3:
                                day, month, year = date_parts
                                if len(year) == 2:
                                    year = '20' + year
                                check_in_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    except:
                        pass
                
                check_out_date = pdf_data.get('departure_date', 'Not Found')
                if (not check_out_date or check_out_date == 'Not Found') and email_check_out:
                    check_out_date = email_check_out
                if check_out_date != 'Not Found' and check_out_date:
                    try:
                        if '/' in str(check_out_date):
                            date_parts = str(check_out_date).split('/')
                            if len(date_parts) == 3:
                                day, month, year = date_parts
                                if len(year) == 2:
                                    year = '20' + year
                                check_out_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    except:
                        pass
                total_amount = pdf_data.get('total_amount', 'Not Found')
                gst_number = pdf_data.get('gst_number', 'Not Found')
                # Use date-formatted filename for Excel
                original_filename = pdf_data.get('filename', '')
                processing_date_for_filename = datetime.now().strftime('%Y%m%d')
                safe_filename = re.sub(r'[^\w\s\-\.]', '', original_filename).strip().replace(' ', '_')
                pdf_filename = f"{processing_date_for_filename}_{safe_filename}"
                drive_file_id = pdf_data.get('drive_file_id', '')
                urgency_level = invoice.get('urgency_level', 'NORMAL')
                
                # Get next S.No
                next_s_no = len(df) + len(new_rows) + 1
                
                new_row = {
                    'S.No': next_s_no,
                    'Processing Date': processing_date,
                    'Email From': email_from,
                    'Email Subject': email_subject,
                    'Assigned To': assigned_folder,
                    'Hotel Name': hotel_name,
                    'Guest Name': guest_name,
                    'Booking Code': booking_code,
                    'Bill Number': bill_number,
                    'Bill Date': bill_date,
                    'Room Number': room_number,
                    'Number of Guests': number_of_guests,
                    'Check-in Date': check_in_date,
                    'Check-out Date': check_out_date,
                    'Total Amount': total_amount,
                    'GST Number': gst_number,
                    'PDF Filename': pdf_filename,
                    'Drive File ID': drive_file_id,
                    'Urgency Level': urgency_level
                }
                new_rows.append(new_row)
        
        if not new_rows:
            print("⚠️  No valid invoice data to add to Excel")
            return None
        
        # Append new data to existing DataFrame
        new_df = pd.DataFrame(new_rows)
        updated_df = pd.concat([df, new_df], ignore_index=True)

        # Ensure legacy columns are removed from the updated dataset
        columns_to_remove = ['PAN Number', 'Amounts Found']
        existing_columns_to_drop = [col for col in columns_to_remove if col in updated_df.columns]
        if existing_columns_to_drop:
            updated_df = updated_df.drop(columns=existing_columns_to_drop)

        desired_column_order = [
            'S.No', 'Processing Date', 'Email From', 'Email Subject', 'Assigned To',
            'Hotel Name', 'Guest Name', 'Booking Code', 'Bill Number', 'Bill Date', 'Room Number',
            'Number of Guests', 'Check-in Date', 'Check-out Date', 'Total Amount',
            'GST Number', 'PDF Filename', 'Drive File ID', 'Urgency Level'
        ]
        available_columns = [col for col in desired_column_order if col in updated_df.columns]
        updated_df = updated_df[available_columns]
        
        # Save updated file with improved formatting
        temp_updated_file = 'temp_master_updated.xlsx'
        with pd.ExcelWriter(temp_updated_file, engine='openpyxl') as writer:
            updated_df.to_excel(writer, sheet_name='Invoice Data', index=False)
            
            # Get the workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Invoice Data']
            
            # Set column widths for better readability
            column_widths = {
                'A': 8,   # S.No
                'B': 20,  # Processing Date
                'C': 25,  # Email From
                'D': 40,  # Email Subject
                'E': 15,  # Assigned To
                'F': 35,  # Hotel Name
                'G': 25,  # Guest Name
                'H': 18,  # Booking Code
                'I': 18,  # Bill Number
                'J': 15,  # Bill Date
                'K': 12,  # Room Number
                'L': 15,  # Number of Guests
                'M': 15,  # Check-in Date
                'N': 15,  # Check-out Date
                'O': 18,  # Total Amount
                'P': 20,  # GST Number
                'Q': 25,  # PDF Filename
                'R': 25,  # Drive File ID
                'S': 12   # Urgency Level
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Add professional formatting
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Apply header formatting
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply borders to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Data rows
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Freeze the header row
            worksheet.freeze_panes = "A2"
            
            # Auto-fit row heights
            for row in worksheet.iter_rows():
                max_height = 0
                for cell in row:
                    if cell.value:
                        # Calculate approximate height based on content
                        lines = str(cell.value).count('\n') + 1
                        height = max(15, lines * 15)  # Minimum 15, 15 per line
                        max_height = max(max_height, height)
                if max_height > 0:
                    worksheet.row_dimensions[row[0].row].height = min(max_height, 60)  # Max 60
            
            # Format date columns properly
            from openpyxl.styles import NamedStyle
            date_style = NamedStyle(name="date_style")
            date_style.number_format = 'YYYY-MM-DD'
            
            # Apply date formatting to date columns (B, I, L, M)
            date_columns = ['B', 'J', 'M', 'N']  # Processing Date, Bill Date, Check-in, Check-out
            for col in date_columns:
                for row in range(2, worksheet.max_row + 1):  # Skip header row
                    cell = worksheet[f'{col}{row}']
                    if cell.value and str(cell.value) != 'Not Found':
                        try:
                            # Try to format as date if it looks like a date
                            cell_value = str(cell.value)
                            if '-' in cell_value or '/' in cell_value:
                                cell.number_format = 'YYYY-MM-DD'
                        except:
                            pass
        
        print(f"📊 Updated Excel with {len(new_rows)} new rows")
        
        # Ensure the ExcelWriter is fully closed before uploading
        # The 'with' statement should handle this, but add explicit wait for Windows
        import time
        time.sleep(0.5)  # Brief wait to ensure file is fully written and closed
        
        # Upload back to Google Drive
        from googleapiclient.http import MediaFileUpload
        media_body = MediaFileUpload(temp_updated_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        updated_file = drive_uploader.service.files().update(
            fileId=master_file_id,
            media_body=media_body
        ).execute()
        
        # Get shareable link
        file_info = drive_uploader.service.files().get(fileId=master_file_id, fields='webViewLink,name').execute()
        shareable_link = file_info.get('webViewLink', f"https://docs.google.com/spreadsheets/d/{master_file_id}/edit")
        
        print(f"📊 Successfully updated master Excel file in Google Drive")
        print(f"📄 File: {file_info.get('name', 'Invoice Processor Master Data.xlsx')}")
        print(f"🔗 View Link: {shareable_link}")
        print(f"🔗 Direct Edit Link: https://docs.google.com/spreadsheets/d/{master_file_id}/edit")
        print(f"📋 File ID: {master_file_id}")
        print(f"\n✅ You can view all uploaded invoice data at: {shareable_link}")
        
        # Clean up temp files - use try/except to handle Windows file locking issues
        temp_files = ['temp_master.xlsx', temp_updated_file]
        for temp_file in temp_files:
            if os.path.exists(temp_file):
                try:
                    # Add a small delay before deletion
                    time.sleep(0.3)
                    os.remove(temp_file)
                except PermissionError as e:
                    # File might be locked - not critical since upload succeeded
                    print(f"⚠️  Could not delete temp file {temp_file} (may be in use): {e}")
                    print(f"   💡 You can manually delete it later - file was successfully uploaded to Drive")
                except Exception as e:
                    print(f"⚠️  Error deleting temp file {temp_file}: {e}")
        
        return shareable_link
        
    except Exception as e:
        print(f"❌ Error updating Excel file: {e}")
        return None

def process_invoices():
    """Main function to process invoices from Gmail"""
    print("📧 FINAL INVOICE PROCESSOR - COMPLETE FLOW")
    print("=" * 50)
    print("1. 📧 Check unread emails")
    print("2. 🤖 Process invoices with AWS Bedrock")
    print("3. ☁️  Upload PDFs to Google Drive folders")
    print("4. 📊 Update Excel sheet with extracted data")
    print("=" * 50)
    
    # Initialize AWS Bedrock extractor
    openai_extractor = OpenAIPropertyExtractor()
    if openai_extractor.enabled:
        print(f"🤖 AWS Bedrock enabled for PDF analysis (Model: {openai_extractor.bedrock_model})")
    else:
        print("⚠️  AWS Bedrock disabled - using text analysis only")
    
    # Initialize Google Drive uploader
    drive_uploader = None
    if os.getenv('ENABLE_GOOGLE_DRIVE_UPLOAD', 'false').lower() == 'true':
        try:
            drive_uploader = GoogleDriveUploader()
            if drive_uploader.authenticate():
                drive_uploader.setup_folders()
                print("✅ Google Drive initialized successfully")
            else:
                print("❌ Google Drive authentication failed")
                return
        except Exception as e:
            print(f"❌ Google Drive setup failed: {e}")
            return
    else:
        print("❌ Google Drive upload disabled")
        return
    
    # Load vendor reference data
    vendor_reference = load_vendor_reference()
    print(f"✅ Loaded {len(vendor_reference)} vendor references")
    
    # Connect to Gmail
    try:
        mail = imaplib.IMAP4_SSL(os.getenv('IMAP_SERVER'), int(os.getenv('IMAP_PORT')))
        mail.login(os.getenv('GMAIL_EMAIL'), os.getenv('GMAIL_PASSWORD'))
        mail.select('inbox')
        print("✅ Connected to Gmail successfully!")
    except Exception as e:
        print(f"❌ Gmail connection failed: {e}")
        return
    
    # Search for unread emails
    try:
        status, messages = mail.search(None, 'UNSEEN')
        unread_emails = messages[0].split()
        print(f"📧 Unread emails: {len(unread_emails)}")
        
        # Get emails from last N days - use a more robust approach
        days_to_search = int(os.getenv('DAYS_TO_SEARCH', '7'))
        
        # Use a fixed reference date to avoid system date issues
        current_date = datetime.now()
        if current_date.year > 2024:  # If system date seems wrong, use a fixed date
            reference_date = datetime(2024, 10, 6)  # Use a known good date
            print(f"⚠️  System date appears incorrect ({current_date.strftime('%Y-%m-%d')}), using reference date")
        else:
            reference_date = current_date
            
        since_date = (reference_date - timedelta(days=days_to_search)).strftime('%d-%b-%Y')
        print(f"🔍 Searching for unread emails since: {since_date}")
        
        # Try different date formats if the first one doesn't work
        try:
            status, messages = mail.search(None, f'UNSEEN SINCE {since_date}')
            recent_emails = messages[0].split()
            print(f"📧 Found {len(recent_emails)} unread emails since {since_date}")
        except Exception as e:
            print(f"⚠️  First date format failed: {e}")
            # Try alternative date format
            since_date_alt = (reference_date - timedelta(days=days_to_search)).strftime('%d-%b-%Y')
            try:
                status, messages = mail.search(None, f'UNSEEN SINCE {since_date_alt}')
                recent_emails = messages[0].split()
                print(f"📧 Found {len(recent_emails)} unread emails since {since_date_alt}")
            except Exception as e2:
                print(f"⚠️  Alternative date format also failed: {e2}")
                # Fallback to just UNSEEN without date filter
                status, messages = mail.search(None, 'UNSEEN')
                recent_emails = messages[0].split()
                print(f"📧 Fallback: Found {len(recent_emails)} unread emails (no date filter)")
        
    except Exception as e:
        print(f"❌ Error searching emails: {e}")
        mail.close()
        mail.logout()
        return
    
    # Process emails
    processed_invoice_ids = []
    invoice_count = 0
    folder_counts = {'KUMAR': 0, 'LAKSHMI': 0, 'MOKSHITHA': 0, 'SANDHYA': 0, 'UNASSIGNED': 0}
    all_invoice_data = []  # Store all invoice data for Excel
    
    for i, email_id in enumerate(recent_emails):
        try:
            print(f"\n📨 Processing email {i+1}/{len(recent_emails)}")
            
            # Fetch email
            status, msg_data = mail.fetch(email_id, '(RFC822)')
            email_message = email.message_from_bytes(msg_data[0][1])
            
            # Extract basic email info
            subject = decode_email_subject(email_message.get('Subject', ''))
            from_email = email_message.get('From', '')
            body = ""
            
            if email_message.is_multipart():
                for part in email_message.walk():
                    if part.get_content_type() == "text/plain":
                        body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                        break
            else:
                body = email_message.get_payload(decode=True).decode('utf-8', errors='ignore')
            
            print(f"   From: {from_email}")
            print(f"   Subject: {subject}")
            
            # Check if it's an invoice email
            if is_invoice_email(subject, body):
                print("   🎯 INVOICE DETECTED!")
                invoice_count += 1
                
                # Extract comprehensive email data
                comprehensive_data = extract_comprehensive_email_data(email_message, subject, from_email, body, openai_extractor)
                email_details = comprehensive_data.get('email_booking_details', {}) or {}
                raw_pdf_analysis = comprehensive_data.get('pdf_analysis', {})
                if isinstance(raw_pdf_analysis, dict):
                    pdf_analysis_data = {k: v for k, v in raw_pdf_analysis.items() if isinstance(v, dict)}
                else:
                    pdf_analysis_data = {}

                if (not pdf_analysis_data or len(pdf_analysis_data) == 0) and email_details:
                    synthetic_entry = {
                        'hotel_name': email_details.get('property_name') or 'Not Found',
                        'guest_name': email_details.get('guest_name') or email_details.get('client_name') or 'Not Found',
                        'arrival_date': email_details.get('check_in_date') or 'Not Found',
                        'departure_date': email_details.get('check_out_date') or 'Not Found',
                        'number_of_pax': email_details.get('guest_count') or 'Not Found',
                        'bill_number': 'Not Found',
                        'bill_date': 'Not Found',
                        'room_number': 'Not Found',
                        'total_amount': 'Not Found',
                        'gst_number': 'Not Found',
                        'pan_number': 'Not Found',
                        'filename': 'email_content',
                        'payload': None,
                        'generated_from_email': True
                    }
                    pdf_analysis_data = {'email_table_1': synthetic_entry}

                comprehensive_data['pdf_analysis'] = pdf_analysis_data

                # Extract vendor name using email table data or improved method
                email_property_name = comprehensive_data.get('email_property_name')
                pdf_data_for_vision = None
                attachments_list = comprehensive_data.get('attachments', [])
                if attachments_list and len(attachments_list) > 0:
                    # Filter to only dictionary attachments
                    valid_attachments = [a for a in attachments_list if isinstance(a, dict)]
                    for attachment in valid_attachments:
                        if attachment.get('content_type') == 'application/pdf' and attachment.get('payload'):
                            pdf_data_for_vision = attachment['payload']
                            break
                
                vendor_name = None
                if email_property_name and isinstance(email_property_name, str):
                    vendor_name = email_property_name.strip()
                    if vendor_name:
                        print(f"   🏨 Property name from email table: {vendor_name}")
                
                if not vendor_name or vendor_name.lower() in ['unknown', 'unknown vendor', 'not found']:
                    vendor_name = extract_vendor_name_improved(
                        subject, from_email, body, 
                        pdf_analysis_data, 
                        pdf_data_for_vision, 
                        openai_extractor
                    )
                
                # Find the correct column for this vendor
                assigned_column = find_vendor_column(vendor_name, vendor_reference)
                
                print(f"   🏨 Extracted vendor: {vendor_name}")
                print(f"   📁 Assigned to: {assigned_column}")
                print(f"   🔧 DEBUG: assigned_column type: {type(assigned_column)}, value: '{assigned_column}'")
                
                # Extract amounts for display
                amounts = comprehensive_data.get('amounts', [])
                if amounts:
                    print(f"   💰 Amounts found: {amounts}")
                
                # Extract urgency level
                urgency = comprehensive_data.get('urgency_level', 'NORMAL')
                print(f"   ⚡ Urgency: {urgency}")
                
                # Process each PDF invoice
                if pdf_analysis_data:
                    pdf_analysis = pdf_analysis_data
                    pdf_count = 0
                    for pdf_key, pdf_data in pdf_analysis.items():
                        # Ensure pdf_data is a dictionary
                        if not isinstance(pdf_data, dict):
                            print(f"   ⚠️  PDF data item {pdf_key} is not a dictionary, skipping")
                            continue
                        
                        # Skip if PDF couldn't be processed
                        if pdf_data.get('hotel_name') == 'PDF couldn\'t be processed':
                            print(f"   ⚠️  Skipping PDF {pdf_count + 1} - couldn't be processed")
                            continue
                        
                        # Skip upload for synthetic entries generated from email content
                        if pdf_data.get('generated_from_email'):
                            print(f"   📝 Using email-derived booking data (no PDF attachment)")
                            continue
                            
                        pdf_count += 1
                        
                        # Upload PDF to Google Drive
                        if drive_uploader:
                            file_id = upload_pdf_to_drive(pdf_data, assigned_column, drive_uploader)
                            if file_id:
                                pdf_data['drive_file_id'] = file_id
                                folder_counts[assigned_column if assigned_column != 'Not Found' else 'UNASSIGNED'] += 1
                                print(f"   💾 Uploaded PDF {pdf_count}: {pdf_data.get('filename', 'unknown.pdf')}")
                            else:
                                print(f"   ❌ Failed to upload PDF {pdf_count}")
                        else:
                            print(f"   ⚠️  Google Drive not available - PDF not uploaded")
                    
                    if pdf_count > 0:
                        print(f"   ✅ Processed {pdf_count} PDF(s) for: {assigned_column} folder")
                else:
                    print("   ⚠️  No PDF analysis data found")
                
                # Store invoice data for Excel
                invoice_data = {
                    'from_email': from_email,
                    'subject': subject,
                    'assigned_column': assigned_column,
                    'urgency_level': urgency,
                    'amounts': amounts,
                    'pdf_analysis': pdf_analysis_data,
                    'booking_code': comprehensive_data.get('booking_code'),  # Store booking code from email
                    'email_booking_details': comprehensive_data.get('email_booking_details', {}),
                    'email_guest_name': comprehensive_data.get('email_guest_name'),
                    'email_client_name': comprehensive_data.get('email_client_name'),
                    'email_property_name': comprehensive_data.get('email_property_name'),
                    'email_check_in_date': comprehensive_data.get('email_check_in_date'),
                    'email_check_out_date': comprehensive_data.get('email_check_out_date')
                }
                all_invoice_data.append(invoice_data)
                
                # Only mark invoice emails as read
                processed_invoice_ids.append(email_id)
            else:
                print("   📄 Regular email (not an invoice) - leaving unread")
            
        except Exception as e:
            print(f"   ❌ Error processing email {i+1}: {e}")
            continue
    
    # Mark processed emails as read
    if processed_invoice_ids:
        try:
            print(f"\n📧 Marking {len(processed_invoice_ids)} invoice emails as read...")
            for email_id in processed_invoice_ids:
                mail.store(email_id, '+FLAGS', '\\Seen')
            print("✅ Invoice emails marked as read")
        except Exception as e:
            print(f"⚠️  Error marking emails as read: {e}")
    
    # Update Excel sheet with all invoice data
    if all_invoice_data:
        excel_link = update_excel_sheet(all_invoice_data, drive_uploader)
        if excel_link:
            print(f"📊 Successfully updated Excel sheet: {excel_link}")
        else:
            print("❌ Failed to update Excel sheet")
        
        # Match invoice data with Excel sheet and update "Invoice Received" status
        print(f"\n🔗 Matching invoice data with Excel sheet...")
        rows_matched = match_invoice_data_with_excel(all_invoice_data)
        if rows_matched > 0:
            print(f"✅ Matched and updated {rows_matched} row(s) in Excel sheet")
        else:
            print(f"⚠️  No rows matched in Excel sheet")
    else:
        print("⚠️  No invoice data to update in Excel")
    
    # Final statistics
    try:
        status, messages = mail.search(None, 'UNSEEN')
        unread_after = len(messages[0].split()) if messages[0] else 0
        print(f"\n📧 Unread emails after processing: {unread_after}")
        print(f"📊 Invoice emails processed: {len(processed_invoice_ids)}")
        print(f"🎯 Invoices found: {invoice_count}")
        
        # Print summary
        print(f"\n📊 INVOICE SUMMARY:")
        print("-" * 30)
        for folder, count in folder_counts.items():
            if count > 0:
                print(f"📁 {folder}: {count} PDFs uploaded")
        
    except Exception as e:
        print(f"⚠️  Error getting final statistics: {e}")
    
    mail.close()
    mail.logout()

def normalize_column_name(col_name):
    """Normalize column name for matching (same as email_to_excel_mapper.py)"""
    if pd.isna(col_name):
        return ''
    return str(col_name).strip().lower().replace('_', ' ').replace('-', ' ')

def normalize_booking_code_value(value):
    """Normalize booking code values for robust comparison"""
    if pd.isna(value):
        return ''
    text = str(value).strip()
    if not text:
        return ''
    text = text.strip('"\'`')
    text = text.replace(',', '').replace(' ', '')
    if re.fullmatch(r'\d+\.0+', text):
        text = text.split('.')[0]
    else:
        try:
            float_val = float(text)
            if float_val.is_integer():
                text = str(int(float_val))
        except Exception:
            pass
    return text.lower()

def find_matching_column(df, field_name):
    """Find the matching column in DataFrame for a given field (same as email_to_excel_mapper.py)"""
    # Normalize field name for matching
    normalized_field = normalize_column_name(field_name)
    
    # Try exact match first
    for col in df.columns:
        if normalize_column_name(col) == normalized_field:
            return col
    
    # Try partial match
    for col in df.columns:
        col_normalized = normalize_column_name(col)
        if normalized_field in col_normalized or col_normalized in normalized_field:
            return col
    
    # Try fuzzy matching with common variations
    field_variations = {
        'paid on': ['paid', 'payment date', 'date paid'],
        'booking code': ['booking', 'code', 'booking id', 'reference', 'confirmation'],
        'guest name': ['guest', 'name', 'customer'],
        'hotel name': ['hotel', 'property', 'venue'],
        'check-in date': ['check in', 'arrival', 'checkin'],
        'check-out date': ['check out', 'departure', 'checkout'],
        'payment': ['payment', 'amount', 'total', 'paid'],
    }
    
    for col in df.columns:
        col_normalized = normalize_column_name(col)
        for key, variations in field_variations.items():
            if key in normalized_field:
                for variation in variations:
                    if variation in col_normalized:
                        return col
    
    return None

def read_matching_excel_file():
    """Read Excel file from local path for matching (same as email_to_excel_mapper.py)"""
    try:
        print(f"\n📊 Reading matching Excel file: {MATCHING_EXCEL_FILE_PATH}")
        
        # Check if file exists
        if not os.path.exists(MATCHING_EXCEL_FILE_PATH):
            print(f"❌ Excel file not found: {MATCHING_EXCEL_FILE_PATH}")
            print(f"   Please check the file path in your .env file (EXCEL_FILE_PATH)")
            return None
        
        # Check if file is readable
        if not os.access(MATCHING_EXCEL_FILE_PATH, os.R_OK):
            print(f"❌ Excel file is not readable: {MATCHING_EXCEL_FILE_PATH}")
            print(f"   Please check file permissions")
            return None
        
        # Check file size
        file_size = os.path.getsize(MATCHING_EXCEL_FILE_PATH)
        if file_size == 0:
            print(f"❌ Excel file is empty (0 bytes): {MATCHING_EXCEL_FILE_PATH}")
            return None
        
        # Try reading with pandas
        print(f"   📄 File size: {file_size:,} bytes")
        df = pd.read_excel(MATCHING_EXCEL_FILE_PATH, engine='openpyxl', header=0)
        print(f"✅ Read Excel file: {len(df)} rows, {len(df.columns)} columns")
        print(f"📋 Columns: {list(df.columns)}")
        return df
        
    except PermissionError:
        print(f"❌ Permission denied: Cannot read Excel file")
        print(f"   The file might be open in Excel. Please close it and try again.")
        return None
    except FileNotFoundError:
        print(f"❌ Excel file not found: {MATCHING_EXCEL_FILE_PATH}")
        return None
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return None

def read_invoice_master_sheet_from_drive(drive_uploader):
    """Read Invoice Processor Master Data from Google Sheets and extract matching fields"""
    if not drive_uploader or not drive_uploader.service:
        print("❌ Google Drive service not available")
        return None
    
    try:
        print(f"\n📊 Reading Invoice Master Sheet from Google Drive...")
        print(f"   File ID: {INVOICE_MASTER_SHEET_ID}")
        
        # Export Google Sheets as Excel file
        request = drive_uploader.service.files().export_media(
            fileId=INVOICE_MASTER_SHEET_ID,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Download to temporary file
        temp_sheet_file = 'temp_invoice_master.xlsx'
        with open(temp_sheet_file, 'wb') as f:
            f.write(request.execute())
        
        print(f"   ✅ Downloaded Invoice Master Sheet")
        
        # Read Excel file with pandas
        df = pd.read_excel(temp_sheet_file, engine='openpyxl', header=0)
        print(f"   ✅ Read {len(df)} rows from Invoice Master Sheet")
        print(f"   📋 Columns: {list(df.columns)}")
        
        # Extract matching fields from the sheet
        extracted_data_list = []
        
        # Find column mappings
        booking_code_col = None
        guest_name_col = None
        hotel_name_col = None
        checkin_col = None
        checkout_col = None
        
        # Map columns (case-insensitive)
        for col in df.columns:
            col_lower = str(col).strip().lower()
            if 'booking code' in col_lower or 'booking' in col_lower:
                booking_code_col = col
            elif 'guest name' in col_lower or ('guest' in col_lower and 'name' in col_lower):
                guest_name_col = col
            elif ('hotel' in col_lower or 'property' in col_lower) and 'name' in col_lower:
                hotel_name_col = col
            elif 'check-in' in col_lower or 'check in' in col_lower or 'arrival' in col_lower:
                checkin_col = col
            elif 'check-out' in col_lower or 'check out' in col_lower or 'departure' in col_lower:
                checkout_col = col
        
        print(f"\n   🔍 Column mappings found:")
        print(f"      Booking Code: {booking_code_col}")
        print(f"      Guest Name: {guest_name_col}")
        print(f"      Hotel Name: {hotel_name_col}")
        print(f"      Check-In Date: {checkin_col}")
        print(f"      Check-Out Date: {checkout_col}")
        
        # Extract data from each row
        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get(booking_code_col if booking_code_col else '')) and \
               pd.isna(row.get(guest_name_col if guest_name_col else '')):
                continue
            
            entry = {
                'Booking Code': str(row[booking_code_col]).strip() if booking_code_col and not pd.isna(row.get(booking_code_col)) else '',
                'Guest Name': str(row[guest_name_col]).strip() if guest_name_col and not pd.isna(row.get(guest_name_col)) else '',
                'Hotel Name': str(row[hotel_name_col]).strip() if hotel_name_col and not pd.isna(row.get(hotel_name_col)) else '',
                'Check-In Date': str(row[checkin_col]).strip() if checkin_col and not pd.isna(row.get(checkin_col)) else '',
                'Check-Out Date': str(row[checkout_col]).strip() if checkout_col and not pd.isna(row.get(checkout_col)) else ''
            }
            
            # Clean up "Not Found" and "nan" values
            for key, value in entry.items():
                if value.lower() in ['not found', 'nan', 'none', '']:
                    entry[key] = ''
            
            # Only add if we have at least one field
            if any(entry.values()):
                extracted_data_list.append(entry)
        
        print(f"   ✅ Extracted {len(extracted_data_list)} entries from Invoice Master Sheet")
        
        # Clean up temp file
        try:
            import time
            time.sleep(0.3)
            if os.path.exists(temp_sheet_file):
                os.remove(temp_sheet_file)
        except:
            pass
        
        return extracted_data_list
        
    except Exception as e:
        print(f"❌ Error reading Invoice Master Sheet from Google Drive: {e}")
        import traceback
        traceback.print_exc()
        return None

def match_invoice_master_sheet_with_excel(drive_uploader):
    """
    Read Invoice Master Sheet from Google Drive, extract matching fields,
    and match with local Excel file to update "Invoice Received" status.
    """
    # Read Invoice Master Sheet from Google Drive
    extracted_data_list = read_invoice_master_sheet_from_drive(drive_uploader)
    if not extracted_data_list:
        print("⚠️  No data extracted from Invoice Master Sheet - skipping matching")
        return 0
    
    # Read local Excel file
    df = read_matching_excel_file()
    if df is None or df.empty:
        print("⚠️  Excel file is empty or invalid - skipping matching")
        return 0
    
    print(f"\n🔍 Matching Invoice Master Sheet data with local Excel file...")
    print(f"🔍 Excel file has {len(df)} rows and {len(df.columns)} columns")
    print(f"🔍 Excel columns: {list(df.columns)}")
    print(f"🔍 Extracted {len(extracted_data_list)} entries from Invoice Master Sheet")
    
    # Create a mapping of field names to Excel columns
    field_to_column = {}
    for field in MATCHING_FIELDS:
        column = find_matching_column(df, field)
        if column:
            field_to_column[field] = column
            print(f"✅ Mapped '{field}' → '{column}'")
        else:
            print(f"⚠️  No matching column found for '{field}'")
    
    if not field_to_column:
        print("❌ No matching columns found. Cannot match Excel rows.")
        return 0
    
    # Check if 'Invoice Received' column exists, if not add it at the end
    if INVOICE_RECEIVED_COLUMN not in df.columns:
        print(f"➕ Adding new column: '{INVOICE_RECEIVED_COLUMN}'")
        df[INVOICE_RECEIVED_COLUMN] = ''
    else:
        print(f"✅ Column '{INVOICE_RECEIVED_COLUMN}' already exists")
    
    rows_updated = 0
    rows_skipped = 0
    
    # Process each extracted data entry
    for idx, entry in enumerate(extracted_data_list):
        if not entry or not any(entry.values()):
            continue
        
        print(f"\n📋 Processing entry {idx+1}:")
        for field, value in entry.items():
            if value:
                print(f"   {field}: {value}")
        
        matching_row_indices = []
        
        # Step 1: Try matching by Booking Code first (Primary Key)
        if 'Booking Code' in field_to_column:
            booking_code_column = field_to_column['Booking Code']
            booking_code_value_raw = entry.get('Booking Code', '')
            booking_code_value = booking_code_value_raw.strip()
            normalized_entry_code = normalize_booking_code_value(booking_code_value_raw)

            if normalized_entry_code:
                print(f"   🔍 Step 1: Matching by Booking Code: '{booking_code_value}' (normalized: '{normalized_entry_code}')")
                try:
                    normalized_series = df[booking_code_column].apply(normalize_booking_code_value)
                    matching_row_indices = normalized_series[normalized_series == normalized_entry_code].index.tolist()

                    if matching_row_indices:
                        print(f"   ✅ Found {len(matching_row_indices)} row(s) with Booking Code: '{booking_code_value}'")
                    else:
                        sample_values = normalized_series.head(5).tolist()
                        print(f"   ⚠️  No row found with Booking Code: '{booking_code_value}' (normalized search)")
                        print(f"      ℹ️ Sample normalized codes from sheet: {sample_values}")
                except Exception as e:
                    print(f"   ❌ Error matching by Booking Code: {e}")
        
        # Step 2: If no Booking Code match, try matching by Guest Name + Hotel Name + Check-In Date + Check-Out Date
        if not matching_row_indices:
            print(f"   🔍 Step 2: Trying to match by Guest Name + Hotel Name + Check-In Date + Check-Out Date")
            
            required_fields = ['Guest Name', 'Hotel Name', 'Check-In Date', 'Check-Out Date']
            missing_fields = [f for f in required_fields if f not in field_to_column]
            
            if missing_fields:
                print(f"   ⚠️  Cannot match: Missing columns for {missing_fields}")
                continue
            
            # Get values for all three fields
            guest_name_value = entry.get('Guest Name', '').strip()
            hotel_name_value = entry.get('Hotel Name', '').strip()
            checkin_value = entry.get('Check-In Date', '').strip()
            checkout_value = entry.get('Check-Out Date', '').strip()
            
            # All four must be present
            if not guest_name_value or not hotel_name_value or not checkin_value or not checkout_value:
                print(f"   ⚠️  Cannot match: Missing required fields")
                continue
            
            print(f"   🔍 Matching by:")
            print(f"      Guest Name: '{guest_name_value}'")
            print(f"      Hotel Name: '{hotel_name_value}'")
            print(f"      Check-In Date: '{checkin_value}'")
            print(f"      Check-Out Date: '{checkout_value}'")
            
            try:
                guest_col = field_to_column['Guest Name']
                hotel_col = field_to_column['Hotel Name']
                checkin_col = field_to_column['Check-In Date']
                checkout_col = field_to_column['Check-Out Date']
                
                # Normalize date formats for matching
                def normalize_date_for_match(date_str):
                    """Normalize date string for comparison"""
                    if pd.isna(date_str) or not date_str:
                        return '', ''
                    
                    # Handle pandas Timestamp
                    if hasattr(date_str, 'strftime'):
                        try:
                            format1 = date_str.strftime('%d-%b-%Y').lower()
                            format2 = date_str.strftime('%d/%m/%Y').lower()
                            return format1, format2
                        except:
                            date_str_lower = str(date_str).lower()
                            return date_str_lower, date_str_lower
                    
                    date_str = str(date_str).strip().lower()
                    format1 = date_str.replace('/', '-').replace(' ', '-').replace('.', '-')
                    format2 = date_str
                    return format1, format2
                
                # Normalize extracted dates
                checkin_norm, checkin_norm2 = normalize_date_for_match(checkin_value)
                checkout_norm, checkout_norm2 = normalize_date_for_match(checkout_value)
                
                # Case-insensitive matching for guest name
                condition1 = df[guest_col].astype(str).str.strip().str.lower() == guest_name_value.lower()
                if not condition1.any():
                    condition1 = df[guest_col].astype(str).str.strip().str.lower().str.contains(
                        guest_name_value.lower(), na=False, regex=False
                    )
                
                # Date matching - check-in
                condition2_list = []
                for row_idx in df.index:
                    excel_checkin = df.at[row_idx, checkin_col]
                    excel_checkin_norm, excel_checkin_norm2 = normalize_date_for_match(excel_checkin)
                    
                    if (checkin_norm == excel_checkin_norm or checkin_norm == excel_checkin_norm2 or
                        checkin_norm2 == excel_checkin_norm or checkin_norm2 == excel_checkin_norm2):
                        condition2_list.append(True)
                    elif checkin_norm and excel_checkin_norm:
                        checkin_clean = checkin_norm.replace('-', '').replace('/', '').replace(' ', '')
                        excel_clean = excel_checkin_norm.replace('-', '').replace('/', '').replace(' ', '')
                        if checkin_clean in excel_clean or excel_clean in checkin_clean:
                            condition2_list.append(True)
                        else:
                            condition2_list.append(False)
                    else:
                        condition2_list.append(False)
                
                condition2 = pd.Series(condition2_list, index=df.index)
                
                # Date matching - check-out
                condition3_list = []
                for row_idx in df.index:
                    excel_checkout = df.at[row_idx, checkout_col]
                    excel_checkout_norm, excel_checkout_norm2 = normalize_date_for_match(excel_checkout)
                    
                    if (checkout_norm == excel_checkout_norm or checkout_norm == excel_checkout_norm2 or
                        checkout_norm2 == excel_checkout_norm or checkout_norm2 == excel_checkout_norm2):
                        condition3_list.append(True)
                    elif checkout_norm and excel_checkout_norm:
                        checkout_clean = checkout_norm.replace('-', '').replace('/', '').replace(' ', '')
                        excel_clean = excel_checkout_norm.replace('-', '').replace('/', '').replace(' ', '')
                        if checkout_clean in excel_clean or excel_clean in checkout_clean:
                            condition3_list.append(True)
                        else:
                            condition3_list.append(False)
                    else:
                        condition3_list.append(False)
                
                condition3 = pd.Series(condition3_list, index=df.index)
                
                # Hotel name matching
                condition4 = df[hotel_col].astype(str).str.strip().str.lower() == hotel_name_value.lower()
                if not condition4.any():
                    condition4 = df[hotel_col].astype(str).str.strip().str.lower().str.contains(
                        hotel_name_value.lower(), na=False, regex=False
                    )
                
                # All four must match
                combined_condition = condition1 & condition2 & condition3 & condition4
                matching_row_indices = df[combined_condition].index.tolist()
                
                if matching_row_indices:
                    print(f"   ✅ Found {len(matching_row_indices)} row(s) matching all three fields")
                else:
                    print(f"   ⚠️  No row found matching Guest Name + Check-In Date + Check-Out Date")
            except Exception as e:
                print(f"   ❌ Error matching by three fields: {e}")
                continue
        
        # Update matching rows
        if matching_row_indices:
            for row_idx in matching_row_indices:
                try:
                    # Check if already marked as "Received"
                    current_value = str(df.at[row_idx, INVOICE_RECEIVED_COLUMN]).strip()
                    if current_value.lower() == INVOICE_RECEIVED_VALUE.lower():
                        print(f"   ⏭️  Row {row_idx + 1} already marked as '{INVOICE_RECEIVED_VALUE}', skipping")
                        rows_skipped += 1
                        continue
                    
                    # Update the Invoice Received column
                    df.at[row_idx, INVOICE_RECEIVED_COLUMN] = INVOICE_RECEIVED_VALUE
                    rows_updated += 1
                    
                    log_info = ""
                    if 'Booking Code' in field_to_column:
                        bc_val = str(df.at[row_idx, field_to_column['Booking Code']])
                        log_info = f"Booking Code: '{bc_val}'"
                    elif 'Guest Name' in field_to_column:
                        gn_val = str(df.at[row_idx, field_to_column['Guest Name']])
                        log_info = f"Guest Name: '{gn_val}'"
                    
                    print(f"   ✅ Updated row {row_idx + 1} - {log_info} → '{INVOICE_RECEIVED_VALUE}'")
                    
                except Exception as e:
                    print(f"   ❌ Error updating row {row_idx + 1}: {e}")
                    continue
        else:
            print(f"   ❌ No matching row found for this entry")
    
    print(f"\n📊 Matching Summary:")
    print(f"   ✅ Rows updated: {rows_updated}")
    print(f"   ⏭️  Rows skipped (already marked): {rows_skipped}")
    print(f"   📋 Total rows in Excel: {len(df)}")
    
    # Save updated Excel file
    if rows_updated > 0:
        try:
            print(f"\n💾 Saving updated Excel file...")
            print(f"   📁 File path: {MATCHING_EXCEL_FILE_PATH}")
            
            # Check if file is open (try to open it in append mode first)
            try:
                # Try to open the file to check if it's locked
                test_file = open(MATCHING_EXCEL_FILE_PATH, 'r+b')
                test_file.close()
            except PermissionError:
                print(f"   ⚠️  File might be open in Excel. Please close it and try again.")
                print(f"   💡 The file will be saved when you close Excel and run the function again.")
                return rows_updated  # Return the count but don't save yet
            
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Save DataFrame to Excel
            print(f"   💾 Writing DataFrame to Excel...")
            df.to_excel(MATCHING_EXCEL_FILE_PATH, index=False, engine='openpyxl')
            print(f"   ✅ DataFrame saved successfully")
            
            # Load with openpyxl to apply formatting
            wb = load_workbook(MATCHING_EXCEL_FILE_PATH)
            ws = wb.active
            
            # Apply header formatting
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            wb.save(MATCHING_EXCEL_FILE_PATH)
            wb.close()
            
            print(f"✅ Saved updated Excel file: {MATCHING_EXCEL_FILE_PATH}")
            print(f"   📊 Total rows in saved file: {len(df)}")
            print(f"   ✅ Rows marked as 'Received': {len(df[df[INVOICE_RECEIVED_COLUMN].astype(str).str.strip().str.lower() == INVOICE_RECEIVED_VALUE.lower()])}")
        except PermissionError as e:
            print(f"❌ Permission denied: Cannot save Excel file")
            print(f"   The file might be open in Excel. Please close it and try again.")
            print(f"   File path: {MATCHING_EXCEL_FILE_PATH}")
        except Exception as e:
            print(f"⚠️  Error saving Excel file: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"\n⚠️  No rows were updated (rows_updated = {rows_updated})")
        print(f"   This could mean:")
        print(f"   1. No matching rows were found in the Excel file")
        print(f"   2. All matching rows were already marked as 'Received'")
        print(f"   3. Column matching failed (check logs above for column mappings)")
    
    return rows_updated

def match_invoice_data_with_excel(invoice_data_list):
    """
    Match invoice data with Excel sheet and update Invoice Received status.
    Uses the same matching logic as email_to_excel_mapper.py
    
    Matching Priority:
    1. Match by Booking Code (Primary Key)
    2. If no Booking Code match, match by Guest Name + Hotel Name + Check-In Date + Check-Out Date (all four required)
    
    Update Logic:
    - Updates the existing row's "Invoice Received" column with "Received"
    - Skips if already marked as "Received"
    """
    # Read Excel file
    df = read_matching_excel_file()
    if df is None or df.empty:
        print("⚠️  Excel file is empty or invalid - skipping matching")
        return 0
    
    print(f"\n🔍 Matching invoice data with Excel sheet...")
    print(f"🔍 Excel file has {len(df)} rows and {len(df.columns)} columns")
    print(f"🔍 Excel columns: {list(df.columns)}")
    
    # Convert invoice data to matching format
    extracted_data_list = []
    for invoice in invoice_data_list:
        pdf_analysis = invoice.get('pdf_analysis', {})
        if not isinstance(pdf_analysis, dict):
            continue
        
        # Get booking code from email
        booking_code = invoice.get('booking_code', '')
        
        # Process each PDF in the invoice
        for pdf_key, pdf_data in pdf_analysis.items():
            if not isinstance(pdf_data, dict):
                continue
            
            # Skip if PDF couldn't be processed
            if pdf_data.get('hotel_name') == 'PDF couldn\'t be processed':
                continue
            
            # Extract matching fields
            guest_name = pdf_data.get('guest_name', '')
            # Use arrival_date as Check-In Date
            check_in_date = pdf_data.get('arrival_date', '')
            # Use departure_date as Check-Out Date
            check_out_date = pdf_data.get('departure_date', '')
            
            # Format dates properly (convert from DD/MM/YYYY to standard format if needed)
            if check_in_date and check_in_date != 'Not Found':
                try:
                    if '/' in str(check_in_date):
                        date_parts = str(check_in_date).split('/')
                        if len(date_parts) == 3:
                            day, month, year = date_parts
                            if len(year) == 2:
                                year = '20' + year
                            check_in_date = f"{day.zfill(2)}/{month.zfill(2)}/{year}"
                except:
                    pass
            
            if check_out_date and check_out_date != 'Not Found':
                try:
                    if '/' in str(check_out_date):
                        date_parts = str(check_out_date).split('/')
                        if len(date_parts) == 3:
                            day, month, year = date_parts
                            if len(year) == 2:
                                year = '20' + year
                            check_out_date = f"{day.zfill(2)}/{month.zfill(2)}/{year}"
                except:
                    pass
            
            hotel_name = pdf_data.get('hotel_name', '')
            if (not hotel_name or hotel_name in ['Not Found', '']) and invoice.get('email_property_name'):
                hotel_name = invoice.get('email_property_name')
            
            # Create entry for matching
            entry = {
                'Booking Code': booking_code if booking_code and booking_code != 'Not Found' else '',
                'Guest Name': guest_name if guest_name and guest_name != 'Not Found' else '',
                'Hotel Name': hotel_name if hotel_name and hotel_name != 'Not Found' else '',
                'Check-In Date': check_in_date if check_in_date and check_in_date != 'Not Found' else '',
                'Check-Out Date': check_out_date if check_out_date and check_out_date != 'Not Found' else ''
            }
            
            # Only add if we have at least one field
            if any(entry.values()):
                extracted_data_list.append(entry)
    
    if not extracted_data_list:
        print("⚠️  No invoice data to match")
        return 0
    
    print(f"🔍 Extracted {len(extracted_data_list)} entries from invoices for matching")
    
    # Create a mapping of field names to Excel columns
    field_to_column = {}
    for field in MATCHING_FIELDS:
        column = find_matching_column(df, field)
        if column:
            field_to_column[field] = column
            print(f"✅ Mapped '{field}' → '{column}'")
        else:
            print(f"⚠️  No matching column found for '{field}'")
    
    if not field_to_column:
        print("❌ No matching columns found. Cannot match Excel rows.")
        return 0
    
    # Check if 'Invoice Received' column exists, if not add it at the end
    if INVOICE_RECEIVED_COLUMN not in df.columns:
        print(f"➕ Adding new column: '{INVOICE_RECEIVED_COLUMN}'")
        df[INVOICE_RECEIVED_COLUMN] = ''
    else:
        print(f"✅ Column '{INVOICE_RECEIVED_COLUMN}' already exists")
    
    rows_updated = 0
    rows_skipped = 0
    
    # Process each extracted data entry
    for idx, entry in enumerate(extracted_data_list):
        if not entry or not any(entry.values()):
            print(f"⚠️  Entry {idx+1} is empty, skipping")
            continue
        
        print(f"\n📋 Processing entry {idx+1}:")
        for field, value in entry.items():
            if value:
                print(f"   {field}: {value}")
        
        matching_row_indices = []
        
        # Step 1: Try matching by Booking Code first (Primary Key)
        if 'Booking Code' in field_to_column:
            booking_code_column = field_to_column['Booking Code']
            booking_code_value_raw = entry.get('Booking Code', '')
            booking_code_value = booking_code_value_raw.strip()
            normalized_entry_code = normalize_booking_code_value(booking_code_value_raw)
            
            if normalized_entry_code:
                print(f"   🔍 Step 1: Matching by Booking Code: '{booking_code_value}'")
                try:
                    normalized_series = df[booking_code_column].apply(normalize_booking_code_value)
                    matching_row_indices = normalized_series[normalized_series == normalized_entry_code].index.tolist()
                    
                    if matching_row_indices:
                        print(f"   ✅ Found {len(matching_row_indices)} row(s) with Booking Code: '{booking_code_value}'")
                    else:
                        sample_values = normalized_series.head(5).tolist()
                        print(f"   ⚠️  No row found with Booking Code: '{booking_code_value}' (normalized search)")
                        print(f"      ℹ️ Sample normalized codes from sheet: {sample_values}")
                except Exception as e:
                    print(f"   ❌ Error matching by Booking Code: {e}")
        
        # Step 2: If no Booking Code match, try matching by Guest Name + Hotel Name + Check-In Date + Check-Out Date
        if not matching_row_indices:
            print(f"   🔍 Step 2: Trying to match by Guest Name + Hotel Name + Check-In Date + Check-Out Date")
            
            required_fields = ['Guest Name', 'Hotel Name', 'Check-In Date', 'Check-Out Date']
            missing_fields = [f for f in required_fields if f not in field_to_column]
            
            if missing_fields:
                print(f"   ⚠️  Cannot match: Missing columns for {missing_fields}")
                continue
            
            # Get values for all three fields
            guest_name_value = entry.get('Guest Name', '').strip()
            hotel_name_value = entry.get('Hotel Name', '').strip()
            checkin_value = entry.get('Check-In Date', '').strip()
            checkout_value = entry.get('Check-Out Date', '').strip()
            
            # All four must be present
            if not guest_name_value or not hotel_name_value or not checkin_value or not checkout_value:
                print(f"   ⚠️  Cannot match: Missing required fields")
                continue
            
            print(f"   🔍 Matching by:")
            print(f"      Guest Name: '{guest_name_value}'")
            print(f"      Hotel Name: '{hotel_name_value}'")
            print(f"      Check-In Date: '{checkin_value}'")
            print(f"      Check-Out Date: '{checkout_value}'")
            
            try:
                guest_col = field_to_column['Guest Name']
                hotel_col = field_to_column['Hotel Name']
                checkin_col = field_to_column['Check-In Date']
                checkout_col = field_to_column['Check-Out Date']
                
                # Normalize date formats for matching
                def normalize_date_for_match(date_str):
                    """Normalize date string for comparison"""
                    if pd.isna(date_str) or not date_str:
                        return '', ''
                    
                    # Handle pandas Timestamp
                    if hasattr(date_str, 'strftime'):
                        try:
                            format1 = date_str.strftime('%d-%b-%Y').lower()
                            format2 = date_str.strftime('%d/%m/%Y').lower()
                            return format1, format2
                        except:
                            date_str_lower = str(date_str).lower()
                            return date_str_lower, date_str_lower
                    
                    date_str = str(date_str).strip().lower()
                    format1 = date_str.replace('/', '-').replace(' ', '-').replace('.', '-')
                    format2 = date_str
                    return format1, format2
                
                # Normalize extracted dates
                checkin_norm, checkin_norm2 = normalize_date_for_match(checkin_value)
                checkout_norm, checkout_norm2 = normalize_date_for_match(checkout_value)
                
                # Case-insensitive matching for guest name
                condition1 = df[guest_col].astype(str).str.strip().str.lower() == guest_name_value.lower()
                if not condition1.any():
                    condition1 = df[guest_col].astype(str).str.strip().str.lower().str.contains(
                        guest_name_value.lower(), na=False, regex=False
                    )
                
                # Date matching - check-in
                condition2_list = []
                for row_idx in df.index:
                    excel_checkin = df.at[row_idx, checkin_col]
                    excel_checkin_norm, excel_checkin_norm2 = normalize_date_for_match(excel_checkin)
                    
                    if (checkin_norm == excel_checkin_norm or checkin_norm == excel_checkin_norm2 or
                        checkin_norm2 == excel_checkin_norm or checkin_norm2 == excel_checkin_norm2):
                        condition2_list.append(True)
                    elif checkin_norm and excel_checkin_norm:
                        checkin_clean = checkin_norm.replace('-', '').replace('/', '').replace(' ', '')
                        excel_clean = excel_checkin_norm.replace('-', '').replace('/', '').replace(' ', '')
                        if checkin_clean in excel_clean or excel_clean in checkin_clean:
                            condition2_list.append(True)
                        else:
                            condition2_list.append(False)
                    else:
                        condition2_list.append(False)
                
                condition2 = pd.Series(condition2_list, index=df.index)
                
                # Date matching - check-out
                condition3_list = []
                for row_idx in df.index:
                    excel_checkout = df.at[row_idx, checkout_col]
                    excel_checkout_norm, excel_checkout_norm2 = normalize_date_for_match(excel_checkout)
                    
                    if (checkout_norm == excel_checkout_norm or checkout_norm == excel_checkout_norm2 or
                        checkout_norm2 == excel_checkout_norm or checkout_norm2 == excel_checkout_norm2):
                        condition3_list.append(True)
                    elif checkout_norm and excel_checkout_norm:
                        checkout_clean = checkout_norm.replace('-', '').replace('/', '').replace(' ', '')
                        excel_clean = excel_checkout_norm.replace('-', '').replace('/', '').replace(' ', '')
                        if checkout_clean in excel_clean or excel_clean in checkout_clean:
                            condition3_list.append(True)
                        else:
                            condition3_list.append(False)
                    else:
                        condition3_list.append(False)
                
                condition3 = pd.Series(condition3_list, index=df.index)
                
                # Hotel name matching
                condition4 = df[hotel_col].astype(str).str.strip().str.lower() == hotel_name_value.lower()
                if not condition4.any():
                    condition4 = df[hotel_col].astype(str).str.strip().str.lower().str.contains(
                        hotel_name_value.lower(), na=False, regex=False
                    )
                
                # All four must match
                combined_condition = condition1 & condition2 & condition3 & condition4
                matching_row_indices = df[combined_condition].index.tolist()
                
                if matching_row_indices:
                    print(f"   ✅ Found {len(matching_row_indices)} row(s) matching all three fields")
                else:
                    print(f"   ⚠️  No row found matching Guest Name + Check-In Date + Check-Out Date")
            except Exception as e:
                print(f"   ❌ Error matching by three fields: {e}")
                continue
        
        # Update matching rows
        if matching_row_indices:
            for row_idx in matching_row_indices:
                try:
                    # Check if already marked as "Received"
                    current_value = str(df.at[row_idx, INVOICE_RECEIVED_COLUMN]).strip()
                    if current_value.lower() == INVOICE_RECEIVED_VALUE.lower():
                        print(f"   ⏭️  Row {row_idx + 1} already marked as '{INVOICE_RECEIVED_VALUE}', skipping")
                        rows_skipped += 1
                        continue
                    
                    # Update the Invoice Received column
                    df.at[row_idx, INVOICE_RECEIVED_COLUMN] = INVOICE_RECEIVED_VALUE
                    rows_updated += 1
                    
                    log_info = ""
                    if 'Booking Code' in field_to_column:
                        bc_val = str(df.at[row_idx, field_to_column['Booking Code']])
                        log_info = f"Booking Code: '{bc_val}'"
                    elif 'Guest Name' in field_to_column:
                        gn_val = str(df.at[row_idx, field_to_column['Guest Name']])
                        log_info = f"Guest Name: '{gn_val}'"
                    
                    print(f"   ✅ Updated row {row_idx + 1} - {log_info} → '{INVOICE_RECEIVED_VALUE}'")
                    
                except Exception as e:
                    print(f"   ❌ Error updating row {row_idx + 1}: {e}")
                    continue
        else:
            print(f"   ❌ No matching row found for this entry")
    
    print(f"\n📊 Matching Summary:")
    print(f"   ✅ Rows updated: {rows_updated}")
    print(f"   ⏭️  Rows skipped (already marked): {rows_skipped}")
    print(f"   📋 Total rows in Excel: {len(df)}")
    
    # Save updated Excel file
    if rows_updated > 0:
        try:
            print(f"\n💾 Saving updated Excel file...")
            print(f"   📁 File path: {MATCHING_EXCEL_FILE_PATH}")
            
            # Check if file is open (try to open it in append mode first)
            try:
                # Try to open the file to check if it's locked
                test_file = open(MATCHING_EXCEL_FILE_PATH, 'r+b')
                test_file.close()
            except PermissionError:
                print(f"   ⚠️  File might be open in Excel. Please close it and try again.")
                print(f"   💡 The file will be saved when you close Excel and run the function again.")
                return rows_updated  # Return the count but don't save yet
            
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Save DataFrame to Excel
            print(f"   💾 Writing DataFrame to Excel...")
            df.to_excel(MATCHING_EXCEL_FILE_PATH, index=False, engine='openpyxl')
            print(f"   ✅ DataFrame saved successfully")
            
            # Load with openpyxl to apply formatting
            wb = load_workbook(MATCHING_EXCEL_FILE_PATH)
            ws = wb.active
            
            # Apply header formatting
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            wb.save(MATCHING_EXCEL_FILE_PATH)
            wb.close()
            
            print(f"✅ Saved updated Excel file: {MATCHING_EXCEL_FILE_PATH}")
            print(f"   📊 Total rows in saved file: {len(df)}")
            print(f"   ✅ Rows marked as 'Received': {len(df[df[INVOICE_RECEIVED_COLUMN].astype(str).str.strip().str.lower() == INVOICE_RECEIVED_VALUE.lower()])}")
        except PermissionError as e:
            print(f"❌ Permission denied: Cannot save Excel file")
            print(f"   The file might be open in Excel. Please close it and try again.")
            print(f"   File path: {MATCHING_EXCEL_FILE_PATH}")
        except Exception as e:
            print(f"⚠️  Error saving Excel file: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"\n⚠️  No rows were updated (rows_updated = {rows_updated})")
        print(f"   This could mean:")
        print(f"   1. No matching rows were found in the Excel file")
        print(f"   2. All matching rows were already marked as 'Received'")
        print(f"   3. Column matching failed (check logs above for column mappings)")
    
    return rows_updated

def match_master_sheet_with_excel():
    """Standalone function to match Invoice Master Sheet with local Excel file"""
    print("=" * 70)
    print("🔗 MATCHING INVOICE MASTER SHEET WITH EXCEL")
    print("=" * 70)
    print("1. Reading Invoice Master Sheet from Google Drive")
    print("2. Extracting matching fields (Booking Code, Guest Name, Dates)")
    print("3. Matching with local Excel file")
    print("4. Updating 'Invoice Received' status")
    print("=" * 70)
    
    # Initialize Google Drive uploader
    drive_uploader = None
    try:
        drive_uploader = GoogleDriveUploader()
        if drive_uploader.authenticate():
            print("✅ Google Drive initialized successfully")
        else:
            print("❌ Google Drive authentication failed")
            return
    except Exception as e:
        print(f"❌ Google Drive setup failed: {e}")
        return
    
    # Match Invoice Master Sheet with local Excel
    rows_matched = match_invoice_master_sheet_with_excel(drive_uploader)
    
    if rows_matched > 0:
        print(f"\n✅ Successfully matched and updated {rows_matched} row(s) in Excel file")
    else:
        print(f"\n⚠️  No rows matched or updated")

if __name__ == "__main__":
    import sys
    # Check if running matching function standalone
    if len(sys.argv) > 1 and sys.argv[1] == '--match':
        match_master_sheet_with_excel()
    else:
        process_invoices()
